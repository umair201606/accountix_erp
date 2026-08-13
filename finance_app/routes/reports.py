from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from flask import Blueprint, render_template, request, jsonify, Response, send_file, redirect, url_for, flash
from flask_login import login_required
from sqlalchemy import text
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from collections import defaultdict
from shared.extensions import db
from shared.formatting import (MONEY_FMT_WESTERN, excel_money_format,
                               format_amount)
from shared.models.ledger import ChartOfAccount, JournalEntry, JournalLine
from shared.models.base import User
from shared.models.company_settings import AccountingPeriod, ReportSettings
# Imported at module scope so the table registers before the lazy
# db.create_all() on the first request.
from shared.models.twcf import (TwcfLine, TWCF_IN, TWCF_OUT,  # noqa: F401
                                TWCF_FREQUENCIES, TWCF_IN_CATEGORIES,
                                TWCF_OUT_CATEGORIES, TWCF_USER_IN_CATEGORIES,
                                TWCF_USER_OUT_CATEGORIES)
from shared.ledger_utils import posting_account
from shared.tenancy import scoped_get

finance_bp = Blueprint("finance", __name__, url_prefix="/finance")

ACCOUNT_TYPES = {"asset": "Asset", "liability": "Liability",
                 "equity": "Equity", "revenue": "Revenue", "expense": "Expense"}

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=10, color="555555")
DATA_FONT = Font(size=10)
BOLD_FONT = Font(bold=True, size=10)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def _parse_date(d):
    if not d:
        return None
    if isinstance(d, date):
        return d
    for fmt in ("%Y-%m-%d", "%d/%m/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(d, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _default_period():
    """The period to default reports to: the one containing today (preferring an
    active one), else the most recent active period. Robust even when several
    periods are erroneously marked active at once."""
    today = date.today()
    p = (AccountingPeriod.query
         .filter(AccountingPeriod.start_date <= today,
                 AccountingPeriod.end_date >= today)
         .order_by(AccountingPeriod.is_active.desc(),
                   AccountingPeriod.start_date.desc())
         .first())
    if p:
        return p
    return (AccountingPeriod.query.filter_by(is_active=True)
            .order_by(AccountingPeriod.start_date.desc()).first())


def _resolve_period():
    filter_mode = request.args.get("filter_mode", "period")
    period_id = request.args.get("period_id", type=int)
    from_str = request.args.get("from", "").strip()
    to_str = request.args.get("to", "").strip()
    from_date = _parse_date(from_str) if from_str else None
    to_date = _parse_date(to_str) if to_str else None

    periods = AccountingPeriod.query.order_by(AccountingPeriod.start_date.desc()).all()
    # Also include periods sorted ascending for comparative display
    periods_asc = AccountingPeriod.query.order_by(AccountingPeriod.start_date.asc()).all()
    selected_period_id = period_id

    # Comparative params
    comp_mode = request.args.get("comp_mode", "")
    comp_period_ids_str = request.args.get("comp_period_ids", "")
    comp_period_ids = []
    if comp_period_ids_str:
        for pid in comp_period_ids_str.split(","):
            pid = pid.strip()
            if pid and pid.isdigit():
                comp_period_ids.append(int(pid))
    comp_periods = AccountingPeriod.query.filter(
        AccountingPeriod.id.in_(comp_period_ids)).order_by(AccountingPeriod.start_date.asc()).all() if comp_period_ids else []

    if filter_mode in ("period", "comparative") and period_id:
        period = scoped_get(AccountingPeriod, period_id)
        if period:
            from_date = period.start_date
            to_date = period.end_date

    # On initial page load (no filter params), leave dates empty so reports
    # show the "Select a period" prompt instead of auto-loading, but still
    # pre-select the current period in the dropdown.
    if not period_id and not from_str and not to_str:
        if not selected_period_id:
            active = _default_period()
            if active:
                selected_period_id = active.id
        return (from_date, to_date, periods, selected_period_id, filter_mode,
                from_str, to_str, comp_mode, comp_periods, comp_period_ids_str)

    if not from_date and not to_date:
        active = _default_period()
        if active:
            from_date = active.start_date
            to_date = active.end_date
            if not selected_period_id:
                selected_period_id = active.id
    elif filter_mode == "period" and not selected_period_id:
        active = _default_period()
        if active:
            selected_period_id = active.id

    return (from_date, to_date, periods, selected_period_id, filter_mode,
            from_str, to_str, comp_mode, comp_periods, comp_period_ids_str)


def _eod(d):
    """Inclusive upper bound for an ``entry_date`` filter.

    ``JournalEntry.entry_date`` is a DateTime carrying a real wall-clock time,
    so comparing it against a bare ``date`` excludes every entry posted on that
    day. Widen the cutoff to the end of the day so the boundary is inclusive.
    """
    if d is None:
        return None
    if isinstance(d, datetime):
        return d
    return datetime.combine(d, datetime.max.time())


def _get_account_balance(account_id, as_of=None):
    q = db.session.query(
        db.func.coalesce(db.func.sum(JournalLine.debit), 0).label("dr"),
        db.func.coalesce(db.func.sum(JournalLine.credit), 0).label("cr"),
    ).join(JournalEntry, JournalLine.journal_entry_id == JournalEntry.id
           ).filter(JournalLine.account_id == account_id,
                    JournalEntry.is_posted == True)
    if as_of:
        q = q.filter(JournalEntry.entry_date <= _eod(as_of))
    row = q.first()
    return Decimal(str(row.dr)), Decimal(str(row.cr))


def _all_account_balances(as_of=None, account_types=None):
    q = db.session.query(
        JournalLine.account_id,
        ChartOfAccount.code,
        ChartOfAccount.name,
        ChartOfAccount.type,
        db.func.coalesce(db.func.sum(JournalLine.debit), 0).label("dr"),
        db.func.coalesce(db.func.sum(JournalLine.credit), 0).label("cr"),
    ).join(JournalEntry, JournalLine.journal_entry_id == JournalEntry.id
           ).join(ChartOfAccount, JournalLine.account_id == ChartOfAccount.id
                  ).filter(JournalEntry.is_posted == True)
    if as_of:
        q = q.filter(JournalEntry.entry_date <= _eod(as_of))
    if account_types:
        q = q.filter(ChartOfAccount.type.in_(account_types))
    q = q.group_by(JournalLine.account_id, ChartOfAccount.code,
                   ChartOfAccount.name, ChartOfAccount.type
                   ).order_by(ChartOfAccount.code)
    return q.all()


def _net_income(as_of=None):
    rev = _all_account_balances(as_of, ["revenue"])
    exp = _all_account_balances(as_of, ["expense"])
    total_rev = sum((r.cr - r.dr) for r in rev) if rev else Decimal("0")
    total_exp = sum((e.dr - e.cr) for e in exp) if exp else Decimal("0")
    return total_rev - total_exp


def _period_movements(from_date=None, to_date=None, types=None):
    """Per-account (dr, cr) sums of posted lines within the period.
    Returns {account_id: (Decimal dr, Decimal cr)}."""
    q = db.session.query(
        JournalLine.account_id,
        db.func.coalesce(db.func.sum(JournalLine.debit), 0).label("dr"),
        db.func.coalesce(db.func.sum(JournalLine.credit), 0).label("cr"),
    ).join(JournalEntry, JournalLine.journal_entry_id == JournalEntry.id
           ).filter(JournalEntry.is_posted == True)
    if from_date:
        q = q.filter(JournalEntry.entry_date >= from_date)
    if to_date:
        q = q.filter(JournalEntry.entry_date <= _eod(to_date))
    if types:
        q = q.join(ChartOfAccount, JournalLine.account_id == ChartOfAccount.id
                   ).filter(ChartOfAccount.type.in_(types))
    return {r.account_id: (Decimal(str(r.dr)), Decimal(str(r.cr)))
            for r in q.group_by(JournalLine.account_id).all()}


def _pl_by_section(from_date, to_date):
    """Group one period's P&L accounts by structure section.

    Every P&L account's contribution to profit is (credit - debit); revenue
    is naturally positive, expenses negative, and contra accounts (sales
    returns, purchase discounts) self-correct without special cases.
    """
    movements = _period_movements(from_date, to_date, ["revenue", "expense", "contra-expense"])
    accounts = {a.id: a for a in ChartOfAccount.query.filter(
        ChartOfAccount.type.in_(["revenue", "expense", "contra-expense"])).all()}

    by_section = defaultdict(list)
    for aid, (dr, cr) in movements.items():
        a = accounts.get(aid)
        if a is None:
            continue
        if a.type == "expense":
            contrib = dr - cr if cr > dr else cr - dr
        else:
            contrib = cr - dr
        if dr == 0 and cr == 0:
            continue
        section = a.effective_pl_section() or (
            "other_income" if a.type == "revenue" else "other_operating")
        by_section[section].append({"code": a.code, "name": a.name, "contrib": contrib})
    return by_section


def _pl_rows(from_date, to_date):
    """Sectioned P&L per ReportSettings.pl_structure.

    Each structure entry's ``negate`` flag only flips the DISPLAY sign so
    expense sections read as positive figures under a "Less: ..." label.

    Returns (render_rows, net_profit). Render row kinds:
    header / account / total / subtotal. Section and subtotal rows carry the
    structure key they came from so a comparative column can line its figures
    up with the right section instead of guessing from position.
    """
    settings = ReportSettings.get()
    detail = settings.pl_detail_rows or 10
    by_section = _pl_by_section(from_date, to_date)

    rows, running = [], Decimal("0")
    for entry in settings.pl_structure():
        if "section" in entry:
            items = by_section.get(entry["section"], [])
            if not items:
                continue
            negate = bool(entry.get("negate"))
            disp = (lambda c: -c) if negate else (lambda c: c)
            total_contrib = sum(i["contrib"] for i in items)
            running += total_contrib
            items.sort(key=lambda i: -abs(i["contrib"]))
            shown, hidden = items[:detail], items[detail:]
            rows.append({"kind": "header", "label": entry["label"],
                         "section": entry["section"]})
            for i in shown:
                rows.append({"kind": "account", "code": i["code"], "name": i["name"],
                             "section": entry["section"],
                             "amount": float(disp(i["contrib"]))})
            if hidden:
                rows.append({"kind": "account", "code": "",
                             "name": f"Others ({len(hidden)} accounts)",
                             "section": entry["section"],
                             "shown_codes": [i["code"] for i in shown],
                             "amount": float(disp(sum(i["contrib"] for i in hidden)))})
            rows.append({"kind": "total", "label": f"Total {entry['label']}",
                         "section": entry["section"],
                         "amount": float(disp(total_contrib))})
        elif "subtotal" in entry:
            rows.append({"kind": "subtotal", "label": entry["label"],
                         "subtotal": entry["subtotal"],
                         "amount": float(running)})
    return rows, float(running)


def _pl_period_lookup(from_date, to_date):
    """Display-signed P&L figures for one period, keyed for comparative lookup.

    Returns (accounts, totals, subtotals): {code: amount}, {section_key:
    total}, {subtotal_key: running}. Running the real structure per period is
    what makes a comparative column mean anything — the previous code summed
    every account in the period into every total row, so "Total Sales" in a
    comparative column showed the period's net profit.
    """
    settings = ReportSettings.get()
    by_section = _pl_by_section(from_date, to_date)

    accounts, totals, subtotals = {}, {}, {}
    running = Decimal("0")
    for entry in settings.pl_structure():
        if "section" in entry:
            items = by_section.get(entry["section"], [])
            negate = bool(entry.get("negate"))
            disp = (lambda c: -c) if negate else (lambda c: c)
            total_contrib = sum((i["contrib"] for i in items), Decimal("0"))
            running += total_contrib
            for i in items:
                accounts[i["code"]] = float(disp(i["contrib"]))
            totals[entry["section"]] = float(disp(total_contrib))
        elif "subtotal" in entry:
            subtotals[entry["subtotal"]] = float(running)
    return accounts, totals, subtotals


def _pl_comp_amount(row, lookup):
    """One comparative period's figure for one already-rendered P&L row."""
    accounts, totals, subtotals = lookup
    if row["kind"] == "total":
        return totals.get(row.get("section"), 0.0)
    if row["kind"] == "subtotal":
        return subtotals.get(row.get("subtotal"), 0.0)
    if row.get("code"):
        return accounts.get(row["code"], 0.0)
    # An "Others (N accounts)" row: which accounts fall into it differs from
    # period to period, so take whatever the section total does not attribute
    # to the accounts listed above it. That keeps the column adding up.
    section = row.get("section")
    shown = sum(accounts.get(c, 0.0) for c in row.get("shown_codes", []))
    return round(totals.get(section, 0.0) - shown, 2)


# Every figure on a finance report is money. Written without a format Excel
# shows the raw float — 1234.5 where the report says 1,234.50 — and negatives
# with a bare minus instead of the brackets an accountant reads. The pattern
# follows the company's number format, so the sheet groups digits the way the
# screen and the PDF do; kept as a module constant for the western default.
MONEY_FMT = MONEY_FMT_WESTERN


def _col_label(d):
    """Heading for a column of figures: the date those figures are stated at.

    Every report has to label its columns the same way, on screen and in both
    exports. They did not. The exports headed the current column "Amount" and
    the comparatives with period names ("FY 2025-2026") while the screen used
    dates, so an exported comparative did not match the report it came from —
    and the P&L headed its current column with the *comparative* period's name.
    """
    return d.strftime("%d %B, %Y") if d else "Amount"


def _period_line(from_date=None, to_date=None, as_of=None):
    """The dated line under a report title, worded as the screen words it."""
    if from_date and to_date:
        return f"For the period {_col_label(from_date)} to {_col_label(to_date)}"
    if as_of:
        return f"As at {_col_label(as_of)}"
    return ""


def _company_name():
    try:
        from shared.models.company_settings import CompanyInfo
        info = CompanyInfo.get()
        return (info.company_name or "").strip() if info else ""
    except Exception:
        return ""


# Rows 1-3 of every exported sheet, then a blank row: content starts at 5.
SHEET_FIRST_ROW = 5


def _write_sheet_heading(ws, ncols, title, from_date=None, to_date=None,
                         as_of=None, period=None):
    """Company / title / period, the same three lines the screen shows.

    The exports used to carry a single squashed line with ISO dates and no
    company — "Profit & Loss (2026-07-01 to 2027-06-30)" against a screen
    reading "HEAT WAVE / Profit & Loss Statement / For the period 01 July,
    2026 to 30 June, 2027". Returns the first row free for content.
    """
    lines = [(_company_name(), SUBTITLE_FONT),
             (title, TITLE_FONT),
             (period or _period_line(from_date, to_date, as_of), SUBTITLE_FONT)]
    for r, (text, font) in enumerate(lines, 1):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(ncols, 1))
        c = ws.cell(row=r, column=1, value=text)
        c.font = font
        c.alignment = Alignment(horizontal="center")
    return SHEET_FIRST_ROW


def _looks_numeric(v):
    """Is this cell a figure, whoever formatted it?

    Several reports hand the PDF builder money already rendered — "1,234.50",
    "(9,876.50)". Judging a column only by its Python type classified those as
    text and left-aligned them, so the decimal points stopped lining up. A
    column is numeric if every value in it reads as a number, formatted or not.
    """
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float, Decimal)):
        return True
    if not isinstance(v, str):
        return False
    s = v.strip().replace(",", "")
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1]
    if s.endswith("%"):
        s = s[:-1]
    if not s:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def _cell_text(v):
    """What Excel will actually display, which is what a column has to be wide
    enough for. Sizing from str(1234.5) budgets six characters for the eight
    that get drawn, and the column shows ###### instead."""
    if isinstance(v, bool) or v is None:
        return "" if v is None else str(v)
    if isinstance(v, (int, float, Decimal)):
        # Measured through the same formatter the sheet is set to use, so a
        # bracketed or Indian-grouped figure gets the width it really needs.
        return format_amount(v)
    return str(v)


def _finish_sheet(ws, ncols, first_row=3):
    money_fmt = excel_money_format()
    """Money format and column widths for a hand-built sheet.

    The five export routes each grew their own copy of the width loop and none
    of them set a number format, so the figures came out unformatted and the
    columns too narrow for them. One pass, applied everywhere.
    """
    for row in ws.iter_rows(min_row=first_row, max_col=ncols):
        for c in row:
            if isinstance(c.value, (int, float, Decimal)) and not isinstance(c.value, bool):
                if c.number_format in (None, "General"):
                    c.number_format = money_fmt
    # A section label written into a merged row spans the whole table, so
    # measuring it would size the narrow Code column to "Less: Selling &
    # Distribution Expenses". Those cells are skipped.
    merged = set()
    for rng in ws.merged_cells.ranges:
        if rng.max_col > rng.min_col:
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    merged.add((r, c))

    for ci in range(1, ncols + 1):
        max_len = 0
        for row in ws.iter_rows(min_row=first_row, min_col=ci, max_col=ci):
            cell = row[0]
            if (cell.row, ci) in merged:
                continue
            max_len = max(max_len, len(_cell_text(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = \
            min(max(max_len + 2, 8), 60)


def _build_excel_wb(title, headers, rows, col_widths=None,
                    bold_rows=None, number_format=None,
                    sheet_title=None, from_date=None, to_date=None, as_of=None,
                    period=None):
    if number_format is None:
        number_format = excel_money_format()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_title or title)[:31]

    hdr_row = _write_sheet_heading(ws, len(headers), title,
                                   from_date=from_date, to_date=to_date,
                                   as_of=as_of, period=period)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=hdr_row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN

    bold_rows = set(bold_rows or ())
    for ri, row in enumerate(rows, hdr_row + 1):
        emphasised = (ri - hdr_row - 1) in bold_rows
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = BOLD_FONT if emphasised else DATA_FONT
            c.border = THIN
            numeric = isinstance(val, (int, float, Decimal))
            c.alignment = RIGHT if numeric else LEFT_ALIGN
            if numeric and number_format:
                c.number_format = number_format

    if col_widths is None:
        col_widths = []
        for ci in range(len(headers)):
            max_len = len(str(headers[ci]))
            for row in rows:
                # Width has to cover the formatted figure, not the raw float.
                cell_val = _cell_text(row[ci]) if ci < len(row) else ""
                max_len = max(max_len, len(cell_val))
            col_widths.append(min(max(max_len + 2, 8), 60))

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# The statement styling the screen uses, in one place so the PDF can copy it
# rather than invent its own. Mirrors finance/layouts/base.html:33-68.
PDF_HEAD_BG = colors.HexColor("#1E293B")      # table header, grand total
PDF_RULE = colors.HexColor("#E2E8F0")         # ordinary row underline
PDF_SECTION_BG = colors.HexColor("#F1F5F9")   # section band
PDF_SECTION_FG = colors.HexColor("#334155")
PDF_SECTION_RULE = colors.HexColor("#CBD5E1")
PDF_TOTAL_BG = colors.HexColor("#FAFAFA")
PDF_TOTAL_RULE = colors.HexColor("#94A3B8")
PDF_PROFIT_BG = colors.HexColor("#EFF6FF")
PDF_NEG_BG = colors.HexColor("#FEF2F2")
PDF_NEG_FG = colors.HexColor("#B91C1C")

# Row kinds a report may tag its rows with. Anything else renders plain.
PDF_ROW_KINDS = {"section", "account", "total", "subtotal", "grand", "spacer", "plain"}


def _build_pdf(title, headers, rows, col_widths=None, bold_rows=None,
               subtitle=None, company=None, row_kinds=None, indent_col=None,
               mono_col=None):
    from reportlab.pdfbase.pdfmetrics import stringWidth
    from reportlab.lib.styles import ParagraphStyle

    ncols = len(headers)
    is_landscape = ncols > 5
    pagesize = landscape(A4) if is_landscape else A4
    buf = BytesIO()

    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        # The page width of the sheet actually in use. A4[0] is the portrait
        # width, so on the landscape reports — anything over five columns —
        # the number was drawn 87mm in from the right, on top of the table.
        canvas.drawRightString(pagesize[0] - 20*mm, 10*mm, f"Page {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(buf, pagesize=pagesize,
                            rightMargin=10*mm, leftMargin=10*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []

    # Company / title / period — the screen's heading block, in the same
    # order and the same wording, centred over the table.
    centred = ParagraphStyle("hdr", parent=styles["Normal"], alignment=1,
                             textColor=colors.HexColor("#555555"))
    if company is None:
        company = _company_name()
    if company:
        elements.append(Paragraph(company, centred))
    elements.append(Paragraph(title, styles["Title"]))
    if subtitle:
        elements.append(Paragraph(subtitle, centred))
    elements.append(Spacer(1, 6*mm))
    elements.append(Paragraph(f"Generated: {datetime.now():%Y-%m-%d %H:%M}",
                              styles["Normal"]))
    elements.append(Spacer(1, 4*mm))

    kinds = list(row_kinds or [])
    kinds += ["plain"] * (len(rows) - len(kinds))

    # A section band spans the table, so its label has to sit in the first
    # cell. Reports write it wherever their column layout puts it; move it.
    body = []
    for row, kind in zip(rows, kinds):
        cells = [format_amount(c) if isinstance(c, (int, float, Decimal))
                 and not isinstance(c, bool) else ("" if c is None else str(c))
                 for c in row]
        if kind == "section":
            label = next((c for c in cells if c.strip()), "")
            cells = [label] + [""] * (len(cells) - 1)
        body.append(cells)
    data = [list(headers)] + body
    available_width = doc.width

    # Which columns hold money. Decided from the values before they were turned
    # into strings, so a numeric column is right-aligned and a text one is not
    # — everything but the first column used to be forced right, which threw
    # account names and descriptions against their column edge.
    numeric_cols = set()
    for ci in range(ncols):
        seen = False
        for row in rows:
            if ci >= len(row) or row[ci] in (None, ""):
                continue
            if not _looks_numeric(row[ci]):
                seen = False
                break
            seen = True
        if seen:
            numeric_cols.add(ci)

    base_font = 8
    header_font_size = base_font + 1
    # Reportlab pads each cell 6pt left and right by default. The old budget of
    # +10pt for a whole column was less than that 12pt, so every column came
    # out narrower than its own widest word and the text broke mid-word —
    # "Vouche/r", "Balanc/e", "2026-07-0/1". Pad explicitly and budget for it.
    pad_x = 4
    if col_widths is None:
        max_widths = [0] * ncols
        for ri, row in enumerate(data):
            # The header row is set larger and bold, so it needs measuring at
            # its own size or a long heading overflows its column.
            fname = "Helvetica-Bold" if ri == 0 else "Helvetica"
            fsize = header_font_size if ri == 0 else base_font
            for ci, val in enumerate(row):
                w = stringWidth(str(val), fname, fsize)
                max_widths[ci] = max(max_widths[ci], w)
        col_widths = [w + 2 * pad_x + 2 for w in max_widths]

    total_width = sum(col_widths)
    if total_width > available_width:
        scale = available_width / total_width
        # Shrink the type with the columns. Scaling the widths alone left the
        # text at its original size inside narrower cells, so it wrapped or ran
        # over the grid instead of fitting.
        base_font = max(5.5, base_font * scale)
        header_font_size = max(6.0, base_font + 1)
        col_widths = [w * scale for w in col_widths]

    font_size = base_font

    cell_style = ParagraphStyle("cell", fontName="Helvetica", fontSize=font_size,
                                leading=font_size * 1.25)
    # A Paragraph draws its own text, so TableStyle's FONTNAME/FONTSIZE/TEXTCOLOR
    # never reached the header — it rendered black, unbolded and at body size on
    # the dark blue fill, which is close to unreadable. The header carries its
    # own style instead.
    head_style = ParagraphStyle("cellhead", fontName="Helvetica-Bold",
                                fontSize=header_font_size,
                                leading=header_font_size * 1.2,
                                textColor=colors.white, alignment=1)
    # A Paragraph draws its own text, so a TableStyle FONTNAME or TEXTCOLOR
    # never reaches it — the same trap the header fell into. Each row kind
    # therefore carries the style its label needs.
    def _kind_style(name, **kw):
        return ParagraphStyle(name, parent=cell_style, **kw)

    label_styles = {
        "section": _kind_style("s", fontName="Helvetica-Bold",
                               textColor=PDF_SECTION_FG),
        "total": _kind_style("t", fontName="Helvetica-Bold"),
        "subtotal": _kind_style("st", fontName="Helvetica-Bold"),
        "subtotal_neg": _kind_style("stn", fontName="Helvetica-Bold",
                                    textColor=PDF_NEG_FG),
        "grand": _kind_style("g", fontName="Helvetica-Bold",
                             textColor=colors.white),
        "mono": _kind_style("m", fontName="Courier"),
    }

    table_data = []
    for ri, row in enumerate(data):
        kind = "header" if ri == 0 else kinds[ri - 1]
        table_row = []
        for ci, val in enumerate(row):
            if ri == 0:
                table_row.append(Paragraph(str(val), head_style))
            elif ci in numeric_cols:
                # Short and right-aligned; a plain string keeps TableStyle's
                # alignment, font and colour in charge.
                table_row.append(val)
            else:
                key = kind
                if kind == "subtotal" and any(str(c).startswith("(") for c in row):
                    key = "subtotal_neg"
                elif kind == "account" and ci == mono_col:
                    key = "mono"
                # Text wraps inside its column. Left as a plain string it would
                # run over into the next cell once the columns were scaled down.
                table_row.append(Paragraph(str(val), label_styles.get(key, cell_style)))
        table_data.append(table_row)

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    pad = 4 if font_size >= 7 else 2
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), PDF_HEAD_BG),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        # Without these the table kept reportlab's 10pt default: the size
        # computed to make the columns fit was only ever applied to the few
        # cells wrapped in a Paragraph, so wide reports ran off the page.
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), header_font_size),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), font_size),
        ("LEADING", (0, 0), (-1, -1), font_size * 1.25),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), pad),
        ("BOTTOMPADDING", (0, 0), (-1, -1), pad),
        ("LEFTPADDING", (0, 0), (-1, -1), pad_x),
        ("RIGHTPADDING", (0, 0), (-1, -1), pad_x),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    # Money right, words left.
    for ci in range(ncols):
        style.append(("ALIGN", (ci, 1), (ci, -1),
                      "RIGHT" if ci in numeric_cols else "LEFT"))

    # A financial statement is ruled horizontally, not boxed. The old blanket
    # GRID plus alternating row fills made every report — sections, totals and
    # account lines alike — read as one undifferentiated spreadsheet dump.
    for ri, kind in enumerate(kinds, 1):
        if kind == "spacer":
            continue
        if kind == "section":
            style += [
                ("SPAN", (0, ri), (-1, ri)),
                ("BACKGROUND", (0, ri), (-1, ri), PDF_SECTION_BG),
                ("TEXTCOLOR", (0, ri), (-1, ri), PDF_SECTION_FG),
                ("FONTNAME", (0, ri), (-1, ri), "Helvetica-Bold"),
                ("LINEABOVE", (0, ri), (-1, ri), 0.5, PDF_SECTION_RULE),
                ("ALIGN", (0, ri), (-1, ri), "LEFT"),
            ]
        elif kind == "total":
            style += [
                ("BACKGROUND", (0, ri), (-1, ri), PDF_TOTAL_BG),
                ("FONTNAME", (0, ri), (-1, ri), "Helvetica-Bold"),
                ("LINEABOVE", (0, ri), (-1, ri), 0.6, PDF_TOTAL_RULE),
            ]
        elif kind == "subtotal":
            negative = any(str(c).startswith("(") for c in body[ri - 1])
            style += [
                ("BACKGROUND", (0, ri), (-1, ri), PDF_NEG_BG if negative else PDF_PROFIT_BG),
                ("FONTNAME", (0, ri), (-1, ri), "Helvetica-Bold"),
                ("LINEABOVE", (0, ri), (-1, ri), 1.0, PDF_HEAD_BG),
                ("LINEBELOW", (0, ri), (-1, ri), 1.0, PDF_HEAD_BG),
            ]
            if negative:
                style.append(("TEXTCOLOR", (0, ri), (-1, ri), PDF_NEG_FG))
        elif kind == "grand":
            style += [
                ("BACKGROUND", (0, ri), (-1, ri), PDF_HEAD_BG),
                ("TEXTCOLOR", (0, ri), (-1, ri), colors.white),
                ("FONTNAME", (0, ri), (-1, ri), "Helvetica-Bold"),
            ]
        else:
            style.append(("LINEBELOW", (0, ri), (-1, ri), 0.4, PDF_RULE))
            if kind == "account":
                if indent_col is not None:
                    style.append(("LEFTPADDING", (indent_col, ri), (indent_col, ri),
                                  pad_x + 10))
                if mono_col is not None:
                    style.append(("FONTNAME", (mono_col, ri), (mono_col, ri), "Courier"))
    for ri in (bold_rows or ()):
        r = ri + 1
        style += [("FONTNAME", (0, r), (-1, r), "Helvetica-Bold"),
                  ("BACKGROUND", (0, r), (-1, r), colors.HexColor("#E8EEF5"))]
    t.setStyle(TableStyle(style))
    elements.append(t)
    # These belong to build(), not to the SimpleDocTemplate constructor, which
    # silently accepted and ignored them — so no finance PDF ever carried a
    # page number.
    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════
# 1. DASHBOARD
# ═══════════════════════════════════════════════

@finance_bp.route("/")
@login_required
def dashboard():
    return render_template("finance/dashboard.html", now=datetime.utcnow())


# ═══════════════════════════════════════════════
# 2. GENERAL LEDGER
# ═══════════════════════════════════════════════

def _get_descendant_ids(account_id):
    ids = [account_id]
    for child in ChartOfAccount.query.filter_by(parent_id=account_id, is_active=True).all():
        ids.extend(_get_descendant_ids(child.id))
    return ids


def _get_leaf_descendant_ids(account_id):
    """Get only leaf (no-children) descendant IDs, excluding the head itself."""
    leaf_ids = []
    for child in ChartOfAccount.query.filter_by(parent_id=account_id, is_active=True).all():
        grand_children = ChartOfAccount.query.filter_by(parent_id=child.id, is_active=True).count()
        if grand_children == 0:
            leaf_ids.append(child.id)
        else:
            leaf_ids.extend(_get_leaf_descendant_ids(child.id))
    return list(set(leaf_ids))


def _get_ledger_sections(account_ids, from_date, to_date):
    """Per-account ledger: opening balance (all posted activity before the
    period), movements during the period with a running balance, and a
    closing balance labelled Dr/Cr."""
    sections = []
    for aid in account_ids:
        account = scoped_get(ChartOfAccount, aid)
        if not account:
            continue
        opening = Decimal("0")
        if from_date:
            odr, ocr = _get_account_balance(aid, from_date - timedelta(days=1))
            opening = odr - ocr
        q = JournalLine.query.join(JournalEntry).filter(
            JournalLine.account_id == aid,
            JournalEntry.is_posted == True,
        )
        if from_date:
            q = q.filter(JournalEntry.entry_date >= from_date)
        if to_date:
            q = q.filter(JournalEntry.entry_date <= _eod(to_date))
        q = q.order_by(JournalEntry.entry_date, JournalEntry.id)
        lines = q.all()
        rows = []
        balance = opening
        total_dr = total_cr = Decimal("0")
        for line in lines:
            dr = Decimal(str(line.debit))
            cr = Decimal(str(line.credit))
            balance += dr - cr
            total_dr += dr
            total_cr += cr
            rows.append({
                "date": line.entry.entry_date.strftime("%Y-%m-%d") if line.entry.entry_date else "",
                "voucher": line.entry.voucher_number or "",
                "description": line.description or line.entry.description or "",
                "debit": float(dr) if dr else 0,
                "credit": float(cr) if cr else 0,
                "balance": float(balance),
            })
        if not rows and opening == 0:
            # Nothing before or during the period — skip empty accounts when
            # rendering "all accounts" so the report stays readable.
            sections.append({"account": account, "rows": rows, "empty": True,
                             "opening": 0.0, "closing": 0.0,
                             "closing_side": "Dr", "opening_side": "Dr",
                             "total_debit": 0.0, "total_credit": 0.0,
                             "subtotal": 0.0})
            continue
        sections.append({
            "account": account,
            "rows": rows,
            "empty": False,
            "opening": float(abs(opening)),
            "opening_side": "Dr" if opening >= 0 else "Cr",
            "total_debit": float(total_dr),
            "total_credit": float(total_cr),
            "closing": float(abs(balance)),
            "closing_side": "Dr" if balance >= 0 else "Cr",
            "subtotal": float(balance),
        })
    return sections


@finance_bp.route("/ledger")
@login_required
def ledger():
    all_accounts = ChartOfAccount.query.filter_by(is_active=True).order_by(ChartOfAccount.code).all()
    child_ids = {r[0] for r in db.session.query(ChartOfAccount.parent_id).filter(
        ChartOfAccount.parent_id.isnot(None)).distinct().all()}
    heads = [a for a in all_accounts if a.parent_id is not None and a.id in child_ids]
    leaf_accounts = [a for a in all_accounts if a.id not in child_ids]

    from_date, to_date, periods, selected_period_id, filter_mode, from_str, to_str, comp_mode, comp_periods, comp_period_ids_str = _resolve_period()

    # Don't auto-calculate on first page load
    if from_date is None:
        return render_template("finance/ledger.html", account_sections=[],
                               heads=heads, leaf_accounts=leaf_accounts,
                               mode="", selection_mode="custom",
                               selected_account_ids="", selected_head_ids="",
                               from_date=None, to_date=None,
                               periods=periods, selected_period_id=selected_period_id,
                               filter_mode="", from_str="", to_str="",
                               comp_mode="", comp_periods=[], comp_period_ids_str="",
                               now=datetime.utcnow())

    mode = request.args.get("mode", "all")
    selection_mode = request.args.get("selection_mode", "custom")
    account_ids_str = request.args.get("account_ids", "")
    head_ids_str = request.args.get("head_ids", "")

    resolved_ids = []
    if mode == "all":
        resolved_ids = [a.id for a in all_accounts]
    elif mode == "select":
        if selection_mode == "custom" and account_ids_str:
            resolved_ids = [int(x) for x in account_ids_str.split(",") if x.strip().isdigit()]
        elif selection_mode == "head" and head_ids_str:
            head_ids = [int(x) for x in head_ids_str.split(",") if x.strip().isdigit()]
            for hid in head_ids:
                resolved_ids.extend(_get_leaf_descendant_ids(hid))
            resolved_ids = list(set(resolved_ids))

    account_sections = _get_ledger_sections(resolved_ids, from_date, to_date) if resolved_ids else []
    account_sections = [s for s in account_sections if not s.get("empty")]

    fmt = request.args.get("format")
    if fmt and account_sections:
        headers = ["Date", "Voucher #", "Description", "Debit", "Credit", "Balance"]
        if fmt == "excel":
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            for sec in account_sections:
                # The balance stays a number and the Dr/Cr side goes in the
                # label. Written as "1,234.50 Dr" it was text: SUM() skipped
                # it, it sorted apart from the figures above it, and it sat
                # left-aligned in an otherwise right-aligned column.
                data = ([["", "", f"Opening Balance ({sec['opening_side']})",
                          "", "", sec["opening"]]] +
                        [[r["date"], r["voucher"], r["description"], r["debit"], r["credit"], r["balance"]]
                         for r in sec["rows"]])
                ws = wb.create_sheet(title=sec["account"].code[:31])
                hdr_row = _write_sheet_heading(
                    ws, 6, f"{sec['account'].code} - {sec['account'].name}",
                    from_date=from_date, to_date=to_date)
                for ci, h in enumerate(headers, 1):
                    c = ws.cell(row=hdr_row, column=ci, value=h)
                    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = THIN
                for ri, row in enumerate(data, hdr_row + 1):
                    for ci, val in enumerate(row, 1):
                        c = ws.cell(row=ri, column=ci, value=val)
                        c.font = DATA_FONT; c.border = THIN
                        c.alignment = RIGHT if isinstance(val, (int, float, Decimal)) else LEFT_ALIGN
                # Both summary rows go through the same cell treatment as the
                # data above them — they were written bare, so the grid stopped
                # at the last transaction and the totals floated outside the
                # table, unbordered and left-aligned.
                tr = hdr_row + 1 + len(data)
                summary = [
                    ["", "", "Period Movement", sec["total_debit"],
                     sec["total_credit"], ""],
                    ["", "", f"Closing Balance ({sec['closing_side']})",
                     "", "", sec["closing"]],
                ]
                for ri, row in enumerate(summary, tr):
                    for ci, val in enumerate(row, 1):
                        c = ws.cell(row=ri, column=ci, value=val)
                        c.font = BOLD_FONT
                        c.border = THIN
                        c.alignment = RIGHT if isinstance(val, (int, float, Decimal)) else LEFT_ALIGN
                _finish_sheet(ws, 6, first_row=hdr_row + 1)
            out = BytesIO(); wb.save(out); out.seek(0)
            return send_file(out, as_attachment=True, download_name="general_ledger.xlsx",
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        if fmt == "pdf":
            all_data = []
            kinds = []
            for sec in account_sections:
                all_data.append([f"{sec['account'].code}  {sec['account'].name}",
                                 "", "", "", "", ""])
                kinds.append("section")
                # Labels belong in the Description column. "Closing Balance"
                # was landing in the Credit column, which made that column
                # text and left-aligned every credit figure under it, while
                # Debit stayed right — the two money columns disagreed.
                all_data.append(["", "", f"Opening Balance ({sec['opening_side']})",
                                 "", "", sec["opening"]])
                kinds.append("total")
                for r in sec["rows"]:
                    all_data.append([r["date"], r["voucher"], r["description"], r["debit"], r["credit"], r["balance"]])
                    kinds.append("account")
                all_data.append(["", "", "Period Movement", sec["total_debit"], sec["total_credit"], ""])
                kinds.append("total")
                all_data.append(["", "", f"Closing Balance ({sec['closing_side']})",
                                 "", "", sec["closing"]])
                kinds.append("grand")
                all_data.append(["", "", "", "", "", ""])
                kinds.append("spacer")
            hdrs = ["Date", "Voucher #", "Description", "Debit", "Credit", "Balance"]
            pdf_out = _build_pdf("General Ledger", hdrs, all_data,
                                 subtitle=_period_line(from_date, to_date),
                                 row_kinds=kinds)
            return send_file(pdf_out, as_attachment=True, download_name="general_ledger.pdf",
                             mimetype="application/pdf")

    return render_template("finance/ledger.html",
                           account_sections=account_sections,
                           heads=heads, leaf_accounts=leaf_accounts,
                           mode=mode, selection_mode=selection_mode,
                           selected_account_ids=account_ids_str,
                           selected_head_ids=head_ids_str,
                           from_date=from_date, to_date=to_date,
                           periods=periods, selected_period_id=selected_period_id,
                           filter_mode=filter_mode, from_str=from_str, to_str=to_str,
                           comp_mode=comp_mode, comp_periods=comp_periods, comp_period_ids_str=comp_period_ids_str,
                           now=datetime.utcnow())


# ═══════════════════════════════════════════════
# 3. TRIAL BALANCE
# ═══════════════════════════════════════════════

@finance_bp.route("/trial-balance")
@login_required
def trial_balance():
    from_date, to_date, periods, selected_period_id, filter_mode, from_str, to_str, comp_mode, comp_periods, comp_period_ids_str = _resolve_period()

    if from_date is None:
        return render_template("finance/trial_balance.html", rows=[],
                               total_dr_opening=0, total_cr_opening=0,
                               total_dr_movement=0, total_cr_movement=0,
                               total_dr_closing=0, total_cr_closing=0,
                               as_of=date.today(), from_date=None,
                               periods=periods, selected_period_id=selected_period_id,
                               filter_mode="", from_str="", to_str="",
                               comp_mode="", comp_periods=[], comp_period_ids_str="",
                               now=datetime.utcnow())

    as_of = to_date or date.today()

    opening_as_of = None
    if from_date:
        opening_as_of = from_date - timedelta(days=1)

    opening_balances = _all_account_balances(opening_as_of) if opening_as_of else []
    closing_balances = _all_account_balances(as_of)

    # Index closing by account code
    closing_map = {}
    for b in closing_balances:
        closing_map[b.code] = b

    opening_map = {}
    for b in opening_balances:
        opening_map[b.code] = b

    all_codes = set(closing_map.keys()) | set(opening_map.keys())

    rows = []
    total_dr_op = Decimal("0")
    total_cr_op = Decimal("0")
    total_dr_mv = Decimal("0")
    total_cr_mv = Decimal("0")
    total_dr_cl = Decimal("0")
    total_cr_cl = Decimal("0")

    for code in sorted(all_codes):
        cb = closing_map.get(code)
        ob = opening_map.get(code)

        dr_op = ob.dr if ob else Decimal("0")
        cr_op = ob.cr if ob else Decimal("0")
        dr_cl = cb.dr if cb else Decimal("0")
        cr_cl = cb.cr if cb else Decimal("0")
        dr_mv = dr_cl - dr_op
        cr_mv = cr_cl - cr_op

        if dr_cl == cr_cl == 0 and dr_op == cr_op == 0:
            continue

        name = (cb or ob).name
        type_ = ACCOUNT_TYPES.get((cb or ob).type, (cb or ob).type)

        total_dr_op += dr_op
        total_cr_op += cr_op
        total_dr_mv += dr_mv
        total_cr_mv += cr_mv
        total_dr_cl += dr_cl
        total_cr_cl += cr_cl

        rows.append({
            "code": code,
            "name": name,
            "type": type_,
            "dr_opening": float(dr_op),
            "cr_opening": float(cr_op),
            "dr_movement": float(dr_mv),
            "cr_movement": float(cr_mv),
            "dr_closing": float(dr_cl),
            "cr_closing": float(cr_cl),
        })

    # Roll-up subtotals per account class (codes start with the class digit:
    # 1 Assets .. 5 Expenses), inserted after each class's rows.
    CLASS_NAMES = {"1": "Assets", "2": "Liabilities", "3": "Equity",
                   "4": "Revenue", "5": "Expenses"}
    grouped = []
    cls_tot = None
    prev_cls = None

    def close_class(g, tot, cls):
        if tot and cls in CLASS_NAMES:
            g.append({"code": "", "name": f"Total {CLASS_NAMES[cls]}", "type": "",
                      "is_subtotal": True, **{k: tot[k] for k in tot}})

    for r in rows:
        cls = (r["code"] or "?")[0]
        if cls != prev_cls:
            close_class(grouped, cls_tot, prev_cls)
            cls_tot = {k: 0.0 for k in ("dr_opening", "cr_opening", "dr_movement",
                                        "cr_movement", "dr_closing", "cr_closing")}
            prev_cls = cls
        grouped.append(r)
        for k in cls_tot:
            cls_tot[k] += r[k]
    close_class(grouped, cls_tot, prev_cls)
    rows = grouped

    # Comparative: compute closing balances for each comparative period
    comp_closing_data = []
    comp_class_totals = []
    if comp_mode and comp_periods:
        for cp in comp_periods:
            cp_balances = _all_account_balances(cp.end_date)
            cp_map = {b.code: b for b in cp_balances}
            comp_closing_data.append(cp_map)
            cp_totals = {"dr": Decimal("0"), "cr": Decimal("0")}
            for code in all_codes:
                cb = cp_map.get(code)
                dr = cb.dr if cb else Decimal("0")
                cr = cb.cr if cb else Decimal("0")
                cp_totals["dr"] += dr
                cp_totals["cr"] += cr
            comp_class_totals.append({"total_dr_closing": float(cp_totals["dr"]),
                                      "total_cr_closing": float(cp_totals["cr"])})

        # Annotate rows with comp_dr_closing / comp_cr_closing
        for r in rows:
            r["comp_dr_closing"] = []
            r["comp_cr_closing"] = []
            for cp_map in comp_closing_data:
                cb = cp_map.get(r["code"])
                dr = float(cb.dr) if cb else 0
                cr = float(cb.cr) if cb else 0
                r["comp_dr_closing"].append(dr)
                r["comp_cr_closing"].append(cr)

    # Count the comparative data actually built, not the periods that happen
    # to be in the query string. Switching the filter back from comparative
    # leaves comp_period_ids populated with comp_mode cleared, which opened
    # unlabelled columns of zeros.
    n_comp = len(comp_periods) if comp_mode else 0
    fmt = request.args.get("format")
    headers = ["Code", "Account", "Type", "Dr Opening", "Cr Opening",
               "Dr Movement", "Cr Movement", "Dr Closing", "Cr Closing"]
    for cp in comp_periods:
        headers += [f"Dr Clsg ({_col_label(cp.end_date)})",
                    f"Cr Clsg ({_col_label(cp.end_date)})"]

    if fmt == "excel":
        data = []
        for r in rows:
            row_data = [r["code"], r["name"], r["type"],
                        r["dr_opening"], r["cr_opening"],
                        r["dr_movement"], r["cr_movement"],
                        r["dr_closing"], r["cr_closing"]]
            if r.get("comp_dr_closing"):
                for i in range(len(comp_periods)):
                    row_data += [r["comp_dr_closing"][i], r["comp_cr_closing"][i]]
            else:
                row_data += [0, 0] * n_comp
            data.append(row_data)
        totals_row = ["", "TOTAL", "",
                      float(total_dr_op), float(total_cr_op),
                      float(total_dr_mv), float(total_cr_mv),
                      float(total_dr_cl), float(total_cr_cl)]
        for ct in comp_class_totals:
            totals_row += [ct["total_dr_closing"], ct["total_cr_closing"]]
        data.append(totals_row)
        wb_out = _build_excel_wb("Trial Balance", headers, data,
                                 sheet_title="Trial Balance",
                                 from_date=from_date, to_date=to_date, as_of=as_of)
        return send_file(wb_out, as_attachment=True, download_name="trial_balance.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    if fmt == "pdf":
        # Figures stay numeric. _build_pdf formats them and decides column
        # alignment from the values, so pre-rendering them here left every
        # money column classified as text and aligned left.
        data = []
        for r in rows:
            row_data = [r["code"], r["name"], r["type"],
                        r["dr_opening"], r["cr_opening"],
                        r["dr_movement"], r["cr_movement"],
                        r["dr_closing"], r["cr_closing"]]
            if r.get("comp_dr_closing"):
                for i in range(len(comp_periods)):
                    row_data += [r["comp_dr_closing"][i], r["comp_cr_closing"][i]]
            else:
                row_data += ["", ""] * n_comp
            data.append(row_data)
        totals_row = ["", "TOTAL", "",
                      float(total_dr_op), float(total_cr_op),
                      float(total_dr_mv), float(total_cr_mv),
                      float(total_dr_cl), float(total_cr_cl)]
        for ct in comp_class_totals:
            totals_row += [format_amount(ct["total_dr_closing"]),
                           format_amount(ct["total_cr_closing"])]
        kinds = ["account"] * len(data)
        data.append(totals_row)
        kinds.append("grand")
        pdf_out = _build_pdf("Trial Balance", headers, data,
                             subtitle=_period_line(from_date, to_date, as_of),
                             row_kinds=kinds, mono_col=0)
        return send_file(pdf_out, as_attachment=True, download_name="trial_balance.pdf",
                         mimetype="application/pdf")

    return render_template("finance/trial_balance.html", rows=rows,
                           total_dr_opening=float(total_dr_op),
                           total_cr_opening=float(total_cr_op),
                           total_dr_movement=float(total_dr_mv),
                           total_cr_movement=float(total_cr_mv),
                           total_dr_closing=float(total_dr_cl),
                           total_cr_closing=float(total_cr_cl),
                           comp_class_totals=comp_class_totals,
                           as_of=as_of, from_date=from_date,
                           periods=periods, selected_period_id=selected_period_id,
                           filter_mode=filter_mode, from_str=from_str, to_str=to_str,
                           comp_mode=comp_mode, comp_periods=comp_periods, comp_period_ids_str=comp_period_ids_str,
                           now=datetime.utcnow())


# ═══════════════════════════════════════════════
# 4. PROFIT & LOSS
# ═══════════════════════════════════════════════

@finance_bp.route("/profit-loss")
@login_required
def profit_loss():
    from_date, to_date, periods, selected_period_id, filter_mode, from_str, to_str, comp_mode, comp_periods, comp_period_ids_str = _resolve_period()

    if from_date is None:
        return render_template("finance/profit_loss.html", pl_rows=[], net_profit=0,
                               from_date=None, to_date=None,
                               periods=periods, selected_period_id=selected_period_id,
                               filter_mode="", from_str="", to_str="",
                               comp_mode="", comp_periods=[], comp_period_ids_str="",
                               now=datetime.utcnow())

    pl_rows, net_profit = _pl_rows(from_date, to_date)

    # Comparative data: each period re-run through the same P&L structure, so
    # a section total compares against that section and not the whole period.
    if comp_mode and comp_periods:
        comp_lookups = [_pl_period_lookup(cp.start_date, cp.end_date)
                        for cp in comp_periods]
        for row in pl_rows:
            if row["kind"] not in ("account", "total", "subtotal"):
                continue
            row["comp_amounts"] = [row.get("amount", 0)] + [
                _pl_comp_amount(row, lk) for lk in comp_lookups]

    fmt = request.args.get("format")
    if fmt == "excel":
        ncols = 3 + len(comp_periods)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "P&L"
        hdr_row = _write_sheet_heading(ws, ncols, "Profit & Loss Statement",
                                      from_date=from_date, to_date=to_date)
        # The sheet had no column headings at all: with comparatives on, every
        # column from D rightwards was an unlabelled wall of figures and there
        # was no way to tell which period was which.
        for ci, h in enumerate(["Code", "Account", _col_label(to_date)] +
                               [_col_label(cp.end_date) for cp in comp_periods], 1):
            c = ws.cell(row=hdr_row, column=ci, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN
        r = hdr_row + 1
        for row in pl_rows:
            if row["kind"] == "header":
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
                c = ws.cell(row=r, column=1, value=row["label"])
                c.font = Font(bold=True, size=12); c.fill = HEADER_FILL
                c.font = Font(bold=True, size=12, color="FFFFFF")
            elif row["kind"] == "account":
                ws.cell(row=r, column=1, value=row["code"]).font = DATA_FONT
                ws.cell(row=r, column=2, value=row["name"]).font = DATA_FONT
                ws.cell(row=r, column=3, value=row["amount"]).font = DATA_FONT; ws.cell(row=r, column=3).alignment = RIGHT
                for ci, cp in enumerate(comp_periods, 4):
                    amt = row.get("comp_amounts", [])[ci - 3] if row.get("comp_amounts") else 0
                    ws.cell(row=r, column=ci, value=amt).font = DATA_FONT; ws.cell(row=r, column=ci).alignment = RIGHT
            elif row["kind"] == "total":
                ws.cell(row=r, column=2, value=row["label"]).font = BOLD_FONT
                ws.cell(row=r, column=3, value=row["amount"]).font = BOLD_FONT; ws.cell(row=r, column=3).alignment = RIGHT
                for ci, cp in enumerate(comp_periods, 4):
                    amt = row.get("comp_amounts", [])[ci - 3] if row.get("comp_amounts") else 0
                    ws.cell(row=r, column=ci, value=amt).font = BOLD_FONT; ws.cell(row=r, column=ci).alignment = RIGHT
            else:
                ws.cell(row=r, column=2, value=row["label"]).font = Font(bold=True, size=12, color="1F4E79")
                ws.cell(row=r, column=3, value=row["amount"]).font = Font(bold=True, size=12, color="1F4E79"); ws.cell(row=r, column=3).alignment = RIGHT
                for ci, cp in enumerate(comp_periods, 4):
                    amt = row.get("comp_amounts", [])[ci - 3] if row.get("comp_amounts") else 0
                    ws.cell(row=r, column=ci, value=amt).font = Font(bold=True, size=12, color="1F4E79"); ws.cell(row=r, column=ci).alignment = RIGHT
                r += 1
            r += 1
        _finish_sheet(ws, ncols, first_row=hdr_row + 1)
        out = BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, as_attachment=True, download_name="profit_loss.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if fmt == "pdf":
        headers = (["Code", "Account / Section", _col_label(to_date)] +
                   [_col_label(cp.end_date) for cp in comp_periods])
        # The row kinds _pl_rows already carries drive the styling, so the PDF
        # reads as the statement it is instead of an undifferentiated table.
        data, kinds = [], []

        def _amounts(row):
            if row.get("comp_amounts"):
                return [format_amount(a) for a in row["comp_amounts"]]
            return [format_amount(row["amount"])] + [""] * len(comp_periods)

        subtotals = [r for r in pl_rows if r["kind"] == "subtotal"]
        last_subtotal = subtotals[-1] if subtotals else None
        for row in pl_rows:
            if row["kind"] == "header":
                data.append([row["label"].upper()] + [""] * (2 + len(comp_periods)))
                kinds.append("section")
            elif row["kind"] == "account":
                data.append([row["code"], row["name"]] + _amounts(row))
                kinds.append("account")
            elif row["kind"] == "total":
                data.append(["", row["label"]] + _amounts(row))
                kinds.append("total")
            else:
                data.append(["", row["label"]] + _amounts(row))
                # The closing profit line is the statement's bottom line.
                kinds.append("grand" if row is last_subtotal else "subtotal")
                data.append([""] * (3 + len(comp_periods)))
                kinds.append("spacer")
        pdf_out = _build_pdf("Profit &amp; Loss Statement", headers, data,
                             subtitle=_period_line(from_date, to_date),
                             row_kinds=kinds, indent_col=1, mono_col=0)
        return send_file(pdf_out, as_attachment=True, download_name="profit_loss.pdf",
                         mimetype="application/pdf")

    return render_template("finance/profit_loss.html", pl_rows=pl_rows,
                           net_profit=net_profit,
                           from_date=from_date, to_date=to_date,
                           periods=periods, selected_period_id=selected_period_id,
                           filter_mode=filter_mode, from_str=from_str, to_str=to_str,
                           comp_mode=comp_mode, comp_periods=comp_periods, comp_period_ids_str=comp_period_ids_str,
                           now=datetime.utcnow())


# ═══════════════════════════════════════════════
# 5. BALANCE SHEET
# ═══════════════════════════════════════════════

def _bs_data(as_of_date, include_ni=True):
    """Compute balance sheet data for a given date. Returns (assets, liabilities, equity,
    total_assets, total_liabilities, total_equity, net_income)."""
    balances = _all_account_balances(as_of_date)
    ni = _net_income(as_of_date) if include_ni else Decimal("0")
    assets, liabilities, equity = [], [], []
    total_assets = total_liabilities = total_equity = Decimal("0")
    for b in balances:
        if b.type == "asset":
            bal = b.dr - b.cr
            if bal != 0:
                assets.append({"code": b.code, "name": b.name, "amount": float(bal)})
                total_assets += bal
        elif b.type == "liability":
            bal = b.cr - b.dr
            if bal != 0:
                liabilities.append({"code": b.code, "name": b.name, "amount": float(bal)})
                total_liabilities += bal
        elif b.type == "equity":
            bal = b.cr - b.dr
            if bal != 0:
                equity.append({"code": b.code, "name": b.name, "amount": float(bal)})
                total_equity += bal
    if include_ni:
        if ni >= 0:
            equity.append({"code": "", "name": "Net Income (Current Period)", "amount": float(ni)})
        else:
            equity.append({"code": "", "name": "Net Loss (Current Period)", "amount": float(ni)})
        total_equity += Decimal(str(ni))
    return assets, liabilities, equity, total_assets, total_liabilities, total_equity, ni


def _merge_multi_period(base_items, comp_items_list, code_key="code", amount_key="amount"):
    """Merge items from base + comparative periods into a list with 'amounts' array."""
    merged = {}
    for pi, items in enumerate([base_items] + comp_items_list):
        for item in items:
            code = item[code_key]
            if code not in merged:
                merged[code] = {"code": code, "name": item["name"],
                                "amounts": [0.0] * (1 + len(comp_items_list))}
            merged[code]["amounts"][pi] = item[amount_key]
    return sorted(merged.values(), key=lambda x: x["code"])


@finance_bp.route("/balance-sheet")
@login_required
def balance_sheet():
    from_date, to_date, periods, selected_period_id, filter_mode, from_str, to_str, comp_mode, comp_periods, comp_period_ids_str = _resolve_period()

    if from_date is None:
        return render_template("finance/balance_sheet.html", assets=[], liabilities=[], equity=[],
                               total_assets=0, total_liabilities=0, total_equity=0,
                               as_of=date.today(),
                               periods=periods, selected_period_id=selected_period_id,
                               filter_mode="", from_str="", to_str="", comp_mode="",
                               comp_periods=[], comp_period_ids_str="", all_periods=[],
                               merged_assets=[], merged_liabilities=[], merged_equity=[],
                               now=datetime.utcnow())

    as_of = to_date or date.today()
    assets, liabilities, equity, total_assets, total_liabilities, total_equity, ni = _bs_data(as_of)

    # Comparative data
    comp_items_list = []
    comp_totals = []
    all_periods = []
    if comp_mode and comp_periods:
        base_period = scoped_get(AccountingPeriod, selected_period_id) if selected_period_id else None
        all_periods = ([base_period] if base_period else []) + list(comp_periods)
        for cp in comp_periods:
            ca, cl, ce, cta, ctl, cte, cni = _bs_data(cp.end_date)
            comp_items_list.append({"assets": ca, "liabilities": cl, "equity": ce})
            comp_totals.append({"total_assets": float(cta), "total_liabilities": float(ctl), "total_equity": float(cte)})
    else:
        base_period = scoped_get(AccountingPeriod, selected_period_id) if selected_period_id else None
        if base_period:
            all_periods = [base_period]

    # Merged rows carry one 'amounts' array per account (base then each
    # comparative). Built before the export branches because the exports need
    # exactly the rows the screen shows — writing a literal 0 into every
    # comparative cell, as the Excel sheet did, is not a comparative report.
    merged_assets = _merge_multi_period(assets, [cl["assets"] for cl in comp_items_list]) if comp_items_list else []
    merged_liabilities = _merge_multi_period(liabilities, [cl["liabilities"] for cl in comp_items_list]) if comp_items_list else []
    merged_equity = _merge_multi_period(equity, [cl["equity"] for cl in comp_items_list]) if comp_items_list else []

    def _export_rows(items, merged):
        """(code, name, [amount per column]) for one section, comparative or not."""
        if merged:
            return [(m["code"], m["name"], list(m["amounts"])) for m in merged]
        return [(i["code"], i["name"], [i["amount"]]) for i in items]

    def _export_totals(base_total, key):
        return [float(base_total)] + [t[key] for t in comp_totals]

    fmt = request.args.get("format")
    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        ncols = 3 + len(comp_periods)
        first_row = _write_sheet_heading(ws, ncols, "Balance Sheet", as_of=as_of)

        def write_section(ws, sr, section_title, rows, total_label, total_vals):
            ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=ncols)
            ws.cell(row=sr, column=1, value=section_title).font = Font(bold=True, size=12)
            hdr = sr + 1
            h_labels = (["Code", "Account", _col_label(as_of)] +
                        [_col_label(cp.end_date) for cp in comp_periods])
            for ci, h in enumerate(h_labels[:ncols], 1):
                c = ws.cell(row=hdr, column=ci, value=h)
                c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN
            for ri, (code, name, amounts) in enumerate(rows, hdr + 1):
                ws.cell(row=ri, column=1, value=code).font = DATA_FONT
                ws.cell(row=ri, column=1).border = THIN
                ws.cell(row=ri, column=2, value=name).font = DATA_FONT
                ws.cell(row=ri, column=2).border = THIN
                for ci, amt in enumerate(amounts[:ncols - 2], 3):
                    ws.cell(row=ri, column=ci, value=float(amt)).font = DATA_FONT
                    ws.cell(row=ri, column=ci).border = THIN
                    ws.cell(row=ri, column=ci).alignment = RIGHT
            tr = hdr + len(rows) + 1
            ws.cell(row=tr, column=2, value=total_label).font = BOLD_FONT
            ws.cell(row=tr, column=2).border = THIN
            for ci, tv in enumerate(total_vals[:ncols - 2], 3):
                ws.cell(row=tr, column=ci, value=float(tv)).font = BOLD_FONT
                ws.cell(row=tr, column=ci).border = THIN
                ws.cell(row=tr, column=ci).alignment = RIGHT
            return tr + 2

        nr = write_section(ws, first_row, "ASSETS", _export_rows(assets, merged_assets),
                           "Total Assets", _export_totals(total_assets, "total_assets"))
        nr = write_section(ws, nr, "LIABILITIES", _export_rows(liabilities, merged_liabilities),
                           "Total Liabilities", _export_totals(total_liabilities, "total_liabilities"))
        nr = write_section(ws, nr, "EQUITY", _export_rows(equity, merged_equity),
                           "Total Equity", _export_totals(total_equity, "total_equity"))

        # Merge the label columns only. Merging the whole row and then writing
        # the figure into column 3 raised "'MergedCell' object attribute
        # 'value' is read-only" — every Balance Sheet Excel export was a 500,
        # with or without a comparative period.
        ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=2)
        ws.cell(row=nr, column=1, value="LIABILITIES + EQUITY").font = Font(bold=True, size=12)
        lte_vals = [float(total_liabilities + total_equity)] + [
            float(t["total_liabilities"] + t["total_equity"]) for t in comp_totals]
        for ci, lte in enumerate(lte_vals[:ncols - 2], 3):
            ws.cell(row=nr, column=ci, value=lte).font = Font(bold=True, size=12)
        _finish_sheet(ws, ncols, first_row=first_row)
        out = BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, as_attachment=True, download_name="balance_sheet.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if fmt == "pdf":
        ncols = 3 + len(comp_periods)
        headers = (["Code", "Account", _col_label(as_of)] +
                   [_col_label(cp.end_date) for cp in comp_periods])
        blank = [""] * ncols
        all_data, kinds = [], []

        def pdf_section(title, rows, total_label, total_vals):
            # The section headings and the three totals were missing outright:
            # the PDF ran assets straight into liabilities with a blank line
            # between and never stated Total Assets.
            all_data.append([title.upper()] + [""] * (ncols - 1))
            kinds.append("section")
            for code, name, amounts in rows:
                all_data.append([code, name] +
                                [format_amount(a) for a in amounts[:ncols - 2]])
                kinds.append("account")
            all_data.append(["", total_label] +
                            [format_amount(t) for t in total_vals[:ncols - 2]])
            kinds.append("total")
            all_data.append(list(blank))
            kinds.append("spacer")

        pdf_section("Assets", _export_rows(assets, merged_assets),
                    "Total Assets", _export_totals(total_assets, "total_assets"))
        pdf_section("Liabilities", _export_rows(liabilities, merged_liabilities),
                    "Total Liabilities", _export_totals(total_liabilities, "total_liabilities"))
        pdf_section("Equity", _export_rows(equity, merged_equity),
                    "Total Equity", _export_totals(total_equity, "total_equity"))
        lte_vals = [float(total_liabilities + total_equity)] + [
            float(t["total_liabilities"] + t["total_equity"]) for t in comp_totals]
        all_data.append(["", "LIABILITIES + EQUITY"] +
                        [format_amount(v) for v in lte_vals[:ncols - 2]])
        kinds.append("grand")
        pdf_out = _build_pdf("Balance Sheet", headers, all_data,
                             subtitle=_period_line(as_of=as_of),
                             row_kinds=kinds, indent_col=1, mono_col=0)
        return send_file(pdf_out, as_attachment=True, download_name="balance_sheet.pdf",
                         mimetype="application/pdf")

    return render_template("finance/balance_sheet.html", assets=assets,
                           liabilities=liabilities, equity=equity,
                           total_assets=float(total_assets),
                           total_liabilities=float(total_liabilities),
                           total_equity=float(total_equity),
                           as_of=as_of,
                           comp_mode=comp_mode, comp_periods=comp_periods,
                           comp_period_ids_str=comp_period_ids_str,
                           comp_totals=comp_totals,
                           all_periods=all_periods,
                           merged_assets=merged_assets,
                           merged_liabilities=merged_liabilities,
                           merged_equity=merged_equity,
                           periods=periods, selected_period_id=selected_period_id,
                           filter_mode=filter_mode, from_str=from_str, to_str=to_str,
                           now=datetime.utcnow())


# ═══════════════════════════════════════════════
# 6. SOCIE
# ═══════════════════════════════════════════════

SOCIE_EPS = Decimal("0.005")


def _socie_paren(v):
    """Accounting presentation for exports: negatives in parentheses.

    A gap is empty rather than a dash: the PDF builder decides a column is
    money by looking at its values, and one "—" among them classified the
    whole column as text and left-aligned every figure in it.
    """
    if v is None:
        return ""
    return format_amount(v)


def _socie_matrix(period_specs):
    """Build an IFRS-style roll-forward Statement of Changes in Equity.

    ``period_specs`` is a chronological list of ``(start, end)`` tuples. The
    result is a component-per-column matrix that reads top-to-bottom as one
    continuous movement: each period's closing balance row *is* the next
    period's opening, so the reader can follow equity across years without
    re-reading a fresh block per period.

    Column values are anchored to the balance sheet: a column's balance at a
    date is that account's ledger balance (cr − dr) at the date, and the
    Retained Earnings column additionally absorbs cumulative net income plus
    the dividend account (a debit balance, so it subtracts itself). Total
    equity per the closing row therefore always equals balance-sheet equity on
    that date. Every movement row is a *period* movement (difference of two
    cutoff balances), never a cumulative figure, and any part of the change a
    named row does not explain is surfaced as an explicit "Other movements"
    line rather than silently dropped — the statement always foots.

    Returns ``(columns, rows)``; ``columns`` are ``{"key", "label"}`` dicts
    ending with the Total column, ``rows`` are render dicts with
    ``kind`` (opening / closing / movement / subtotal / section),
    ``label``, ``indent`` and ``values`` ({column key: float or None}).
    """
    from shared.coa import ROLE_CODES

    if not period_specs:
        return [], []

    accounts = (ChartOfAccount.query.filter_by(type="equity")
                .order_by(ChartOfAccount.code).all())
    by_code = {a.code: a for a in accounts}
    sc_code = "3-01-01-01-0001"
    re_close = by_code.get(ROLE_CODES.get("retained_earnings"))
    re_open = by_code.get("3-02-01-01-0001")
    div_acct = by_code.get(ROLE_CODES.get("dividends"))
    oci_acct = by_code.get(ROLE_CODES.get("other_comprehensive_income"))
    # Retained-earnings family folds into the single RE column; OCI keeps its
    # own reserve column the way a published statement presents it.
    re_family = {a.id for a in (re_open, re_close, div_acct) if a}

    _cache = {}

    def _at(d):
        """(equity balances by account id, cumulative net income) as of d."""
        if d not in _cache:
            _cache[d] = ({b.account_id: b.cr - b.dr
                          for b in _all_account_balances(d, ["equity"])},
                         _net_income(d))
        return _cache[d]

    # Dates the statement measures: every period boundary.
    cut_dates = []
    for ps, pe in period_specs:
        cut_dates += [ps - timedelta(days=1), pe]

    # Component columns: postable equity accounts outside the RE family that
    # actually move or hold a balance somewhere in the reported span. Share
    # capital is always shown — a SOCIE without it reads as incomplete.
    comp_accounts = []
    for a in accounts:
        if not a.is_postable or a.id in re_family:
            continue
        used = any(abs(_at(d)[0].get(a.id, Decimal("0"))) > SOCIE_EPS for d in cut_dates)
        if used or a.code == sc_code:
            comp_accounts.append(a)

    columns = [{"key": f"a{a.id}", "label": a.name} for a in comp_accounts]
    columns.append({"key": "re", "label": "Retained Earnings"})
    columns.append({"key": "total", "label": "Total Equity"})
    oci_key = f"a{oci_acct.id}" if (oci_acct and any(
        c["key"] == f"a{oci_acct.id}" for c in columns)) else None

    def _balances(d):
        bals, ni = _at(d)
        vals = {f"a{a.id}": float(bals.get(a.id, Decimal("0"))) for a in comp_accounts}
        re_val = ni
        for a in (re_open, re_close, div_acct):
            if a:
                re_val += bals.get(a.id, Decimal("0"))
        vals["re"] = float(re_val)
        vals["total"] = float(sum(Decimal(str(v)) for v in vals.values()))
        return vals

    def _mv(acct, d0, d1):
        if not acct:
            return Decimal("0")
        return _at(d1)[0].get(acct.id, Decimal("0")) - _at(d0)[0].get(acct.id, Decimal("0"))

    def _row(kind, label, values, indent=0):
        # The Total column is always recomputed from the component columns so a
        # row can never disagree with the figures printed beside it.
        full = {c["key"]: values.get(c["key"]) for c in columns if c["key"] != "total"}
        full["total"] = sum(v for v in full.values() if v is not None)
        return {"kind": kind, "label": label, "indent": indent, "values": full}

    rows = []
    prev_end = None
    prev_close = None

    for ps, pe in period_specs:
        cutoff = ps - timedelta(days=1)
        opening = _balances(cutoff)
        closing = _balances(pe)
        span_years = (pe - ps).days >= 350
        term = "year" if span_years else "period"
        ended = f"the {term} ended {pe.strftime('%d %B %Y')}"

        # Opening row only where continuity breaks: the first period, a gap
        # between periods, or a restatement that moved the prior closing.
        adjoining = prev_end is not None and prev_end + timedelta(days=1) == ps
        same_balances = prev_close is not None and all(
            abs((prev_close.get(k) or 0) - (opening.get(k) or 0)) < 0.005 for k in opening)
        if not (adjoining and same_balances):
            label = f"Balance as at {ps.strftime('%d %B %Y')}"
            if adjoining and not same_balances:
                label += " (as restated)"
            rows.append(_row("opening", label, opening))

        explained = defaultdict(Decimal)

        # ── Total comprehensive income ────────────────────────────────
        profit = _at(pe)[1] - _at(cutoff)[1]
        oci_mv = _mv(oci_acct, cutoff, pe) if oci_key else Decimal("0")
        tci_rows = []
        if abs(profit) > SOCIE_EPS:
            word = "Profit" if profit >= 0 else "Loss"
            tci_rows.append(_row("movement", f"{word} for {ended}",
                                 {"re": float(profit)}, indent=1))
            explained["re"] += profit
        if abs(oci_mv) > SOCIE_EPS:
            tci_rows.append(_row("movement", f"Other comprehensive income for {ended}",
                                 {oci_key: float(oci_mv)}, indent=1))
            explained[oci_key] += oci_mv
        if tci_rows:
            rows.append({"kind": "section", "label": "Total comprehensive income",
                         "indent": 0, "values": {}})
            rows += tci_rows
            if len(tci_rows) > 1:
                rows.append(_row("subtotal", f"Total comprehensive income for {ended}",
                                 {"re": float(profit), oci_key: float(oci_mv)}, indent=1))

        # ── Transactions with owners ──────────────────────────────────
        owner_rows = []
        for a in comp_accounts:
            if oci_key and a.id == oci_acct.id:
                continue
            mv = _mv(a, cutoff, pe)
            if abs(mv) <= SOCIE_EPS:
                continue
            if a.code.startswith("3-01"):
                label = ("Shares issued during the " + term) if mv > 0 else \
                        ("Capital reduction during the " + term)
            else:
                label = (f"Transfer to {a.name}") if mv > 0 else (f"Transfer from {a.name}")
            owner_rows.append(_row("movement", label, {f"a{a.id}": float(mv)}, indent=1))
            explained[f"a{a.id}"] += mv
        div_mv = _mv(div_acct, cutoff, pe)
        if abs(div_mv) > SOCIE_EPS:
            owner_rows.append(_row("movement", "Dividends declared", {"re": float(div_mv)},
                                   indent=1))
            explained["re"] += div_mv
        if owner_rows:
            rows.append({"kind": "section", "label": "Transactions with owners",
                         "indent": 0, "values": {}})
            rows += owner_rows
            if len(owner_rows) > 1:
                sub = defaultdict(float)
                for r in owner_rows:
                    for k, v in r["values"].items():
                        if k != "total" and v is not None:
                            sub[k] += v
                rows.append(_row("subtotal", "Total transactions with owners", dict(sub),
                                 indent=1))

        # ── Reconciling remainder ─────────────────────────────────────
        # Direct postings to retained earnings, or to a reserve outside the
        # rows above (prior-period adjustments, opening/closing RE transfers).
        residual = {}
        for c in columns:
            k = c["key"]
            if k == "total":
                continue
            delta = Decimal(str(closing[k])) - Decimal(str(opening[k])) - explained[k]
            if abs(delta) > SOCIE_EPS:
                residual[k] = float(delta)
        if residual:
            rows.append({"kind": "section", "label": "Other movements",
                         "indent": 0, "values": {}})
            rows.append(_row("movement", "Prior-period adjustments and transfers",
                             residual, indent=1))

        rows.append(_row("closing", f"Balance as at {pe.strftime('%d %B %Y')}", closing))
        prev_end, prev_close = pe, closing

    if rows:
        rows[-1]["kind"] = "grand"
    return columns, rows


def _socie_period_specs(from_date, to_date, comp_periods):
    """Current window plus the comparatives, de-duplicated and chronological."""
    specs = [(from_date, to_date)] + [(p.start_date, p.end_date) for p in comp_periods]
    seen, out = set(), []
    for ps, pe in specs:
        if not ps or not pe or (ps, pe) in seen:
            continue
        seen.add((ps, pe))
        out.append((ps, pe))
    out.sort(key=lambda s: (s[0], s[1]))
    return out


@finance_bp.route("/socie")
@login_required
def socie():
    from_date, to_date, periods, selected_period_id, filter_mode, from_str, to_str, comp_mode, comp_periods, comp_period_ids_str = _resolve_period()

    base = dict(periods=periods, selected_period_id=selected_period_id,
                comp_mode=comp_mode, comp_periods=comp_periods,
                comp_period_ids_str=comp_period_ids_str, now=datetime.utcnow())

    if from_date is None:
        return render_template("finance/socie.html", socie_columns=[], socie_rows=[],
                               from_date=None, to_date=None,
                               filter_mode="", from_str="", to_str="", **base)

    specs = _socie_period_specs(from_date, to_date,
                                comp_periods if comp_mode else [])
    columns, rows = _socie_matrix(specs)

    fmt = request.args.get("format")
    # No resolvable period means no columns, and the span line below indexes
    # specs[0] and specs[-1]. The on-screen report already guards this; the
    # export used to answer a 500 for a filter the page renders happily.
    if fmt in ("excel", "pdf") and specs:
        headers = [""] + [c["label"] for c in columns]
        data, bold = [], []
        for r in rows:
            if r["kind"] in ("opening", "closing", "grand", "subtotal", "section"):
                bold.append(len(data))
            line = ["    " * r.get("indent", 0) + r["label"]]
            for c in columns:
                v = r["values"].get(c["key"]) if r["values"] else None
                line.append(_socie_paren(v) if fmt == "pdf" else v)
            data.append(line)
        span = (f"For the period {specs[0][0].strftime('%d %B %Y')} to "
                f"{specs[-1][1].strftime('%d %B %Y')}")
        title = "Statement of Changes in Equity"
        if fmt == "excel":
            out = _build_excel_wb(title, headers, data,
                                  bold_rows=bold, sheet_title="SOCIE",
                                  period=span,
                                  number_format=excel_money_format())
            return send_file(out, as_attachment=True, download_name="socie.xlsx",
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        socie_kinds = ["total" if i in bold else "account"
                       for i in range(len(data))]
        out = _build_pdf(title, headers, data, subtitle=span,
                         row_kinds=socie_kinds, indent_col=0)
        return send_file(out, as_attachment=True, download_name="socie.pdf",
                         mimetype="application/pdf")

    return render_template("finance/socie.html",
                           socie_columns=columns, socie_rows=rows,
                           period_specs=specs,
                           from_date=from_date, to_date=to_date,
                           filter_mode=filter_mode, from_str=from_str, to_str=to_str,
                           **base)


# ═══════════════════════════════════════════════
# 7. CASH FLOW STATEMENT
# ═══════════════════════════════════════════════


def _cash_flow_direct(from_date, to_date):
    """Direct-method cash flow: aggregate cash receipts & payments from journals
    involving cash accounts; counterparty accounts determine the activity.
    Returns (op_items, inv_items, fin_items, opening_cash, closing_cash)."""
    opening_cutoff = from_date - timedelta(days=1)
    all_accts = {a.id: a for a in ChartOfAccount.query.all()}
    cash_ids = [a.id for a in all_accts.values()
                if (a.effective_cash_flow_activity() or "") == "cash" and a.level >= 5]
    cash_set = set(cash_ids)
    lines = db.session.query(JournalLine).join(JournalEntry).filter(
        JournalEntry.is_posted == True,
        JournalEntry.entry_date >= from_date,
        JournalEntry.entry_date <= _eod(to_date),
    ).all()
    by_entry = defaultdict(list)
    for ln in lines:
        by_entry[ln.journal_entry_id].append(ln)
    groups = {"operating": defaultdict(float), "investing": defaultdict(float),
              "financing": defaultdict(float)}
    for eid, entry_lines in by_entry.items():
        cash_lines = [ln for ln in entry_lines if ln.account_id in cash_set]
        other_lines = [ln for ln in entry_lines if ln.account_id not in cash_set]
        if not cash_lines:
            continue
        net_cash = sum(float(ln.debit - ln.credit) for ln in cash_lines)
        if abs(net_cash) < 0.005:
            continue
        other_total = sum(float(ln.credit + ln.debit) for ln in other_lines)
        if other_total == 0:
            continue
        for oln in other_lines:
            oa = all_accts.get(oln.account_id)
            if oa is None:
                continue
            weight = float(oln.credit + oln.debit) / other_total
            portion = net_cash * weight
            activity = oa.effective_cash_flow_activity() or "operating"
            if activity == "cash":
                continue
            a = oa
            while a is not None and a.level > 3:
                a = a.parent
            head = a.name if a is not None else oa.name
            groups[activity][head] += portion
    op_items = []
    for name, v in sorted(groups["operating"].items()):
        if v > 0:
            op_items.append((f"Cash received — {name}", v))
        elif v < 0:
            op_items.append((f"Cash paid — {name}", v))
    inv_items = [(f"{'Proceeds from' if v > 0 else 'Purchase of'} {name}", v)
                 for name, v in sorted(groups["investing"].items())]
    fin_items = [(f"{'Proceeds from' if v > 0 else 'Repayment of'} {name}", v)
                  for name, v in sorted(groups["financing"].items())]

    def cash_balance(as_of):
        total = Decimal("0")
        for cid in cash_ids:
            dr, cr = _get_account_balance(cid, as_of)
            total += dr - cr
        return float(total)
    opening_cash = cash_balance(opening_cutoff)
    closing_cash = cash_balance(to_date)
    return op_items, inv_items, fin_items, opening_cash, closing_cash


@finance_bp.route("/cash-flow")
@login_required
def cash_flow():
    """Cash flow statement — indirect or direct method per ReportSettings."""
    from_date, to_date, periods, selected_period_id, filter_mode, from_str, to_str, comp_mode, comp_periods, comp_period_ids_str = _resolve_period()

    if from_date is None:
        return render_template("finance/cash_flow.html", op_items=[], inv_items=[], fin_items=[],
                               net_operating=0, net_investing=0, net_financing=0,
                               net_change=0, opening_cash=0, closing_cash=0,
                               cash_movement=0, method="indirect",
                               from_date=None, to_date=None,
                               periods=periods, selected_period_id=selected_period_id,
                               filter_mode="", from_str="", to_str="",
                               comp_mode="", comp_periods=[], comp_period_ids_str="",
                               comp_item_maps=[],
                               now=datetime.utcnow())

    settings = ReportSettings.get()
    method = settings.cash_flow_method or "indirect"

    opening_cutoff = from_date - timedelta(days=1)
    all_accts = {a.id: a for a in ChartOfAccount.query.all()}

    if method == "direct":
        op_items, inv_items, fin_items, opening_cash, closing_cash = \
            _cash_flow_direct(from_date, to_date)
        net_operating = sum(v for _, v in op_items)
        net_investing = sum(v for _, v in inv_items)
        net_financing = sum(v for _, v in fin_items)
        net_change = net_operating + net_investing + net_financing
        cash_movement = net_change
    else:
        # ── Indirect method ────────────────────────────────────────────────
        pl_moves = _period_movements(from_date, to_date, ["revenue", "expense", "contra-expense"])
        net_profit = 0.0
        for aid, (dr, cr) in pl_moves.items():
            a = all_accts.get(aid)
            if a is None:
                continue
            # Credit less debit, whatever the account type. Expenses used to
            # take a special case, `dr - cr if cr > dr else cr - dr`, whose
            # arms are both negative — it always added -abs(movement), so an
            # expense with a net credit (a refund, rebate or reversal) cut
            # profit instead of raising it, by twice the amount. _net_income()
            # uses this plain form, and the two have to agree or the exported
            # cash flow contradicts the P&L and opening + change stops
            # equalling closing cash.
            net_profit += float(cr - dr)

        bs_moves = _period_movements(from_date, to_date, ["asset", "liability", "equity"])

        def l3_head(acct):
            a = acct
            while a is not None and a.level > 3:
                a = a.parent
            return a.name if a is not None else acct.name

        groups = {"operating": defaultdict(float), "investing": defaultdict(float),
                  "financing": defaultdict(float)}
        cash_movement = 0.0
        for aid, (dr, cr) in bs_moves.items():
            acct = all_accts.get(aid)
            if acct is None:
                continue
            activity = acct.effective_cash_flow_activity() or "operating"
            if activity == "cash":
                cash_movement += float(dr - cr)
                continue
            effect = float(cr - dr)
            if effect:
                groups[activity][l3_head(acct)] += effect

        op_items = [("Net Profit / (Loss) for the period", net_profit)]
        op_items += [(f"(Increase) / Decrease in {name}" if v < 0 else
                      f"Decrease / (Increase) in {name}", v)
                     for name, v in sorted(groups["operating"].items())]
        inv_items = [(f"Movement in {name}", v) for name, v in sorted(groups["investing"].items())]
        fin_items = [(f"Movement in {name}", v) for name, v in sorted(groups["financing"].items())]

        net_operating = sum(v for _, v in op_items)
        net_investing = sum(v for _, v in inv_items)
        net_financing = sum(v for _, v in fin_items)
        net_change = net_operating + net_investing + net_financing

        cash_ids = [a.id for a in all_accts.values()
                    if (a.effective_cash_flow_activity() or "") == "cash" and a.level >= 5]

        def cash_balance(as_of):
            total = Decimal("0")
            for cid in cash_ids:
                dr, cr = _get_account_balance(cid, as_of)
                total += dr - cr
            return float(total)
        opening_cash = cash_balance(opening_cutoff)
        closing_cash = cash_balance(to_date)

    # Comparative: per-item values for each comparative period
    comp_item_maps = []
    if comp_mode and comp_periods:
        # Build lookup: for each comp period, map item-label -> value
        comp_item_maps = []  # list of {label: val} per period
        for cp in comp_periods:
            cp_from = cp.start_date
            cp_to = cp.end_date
            cmp_map = {}
            if method == "direct":
                cp_op, cp_inv, cp_fin, cp_oc, cp_cc = _cash_flow_direct(cp_from, cp_to)
                for item_list in [cp_op, cp_inv, cp_fin]:
                    for name, val in item_list:
                        cmp_map[name] = val
                cmp_map["__net_op__"] = sum(v for _, v in cp_op)
                cmp_map["__net_inv__"] = sum(v for _, v in cp_inv)
                cmp_map["__net_fin__"] = sum(v for _, v in cp_fin)
                cmp_map["__net_chg__"] = cmp_map["__net_op__"] + cmp_map["__net_inv__"] + cmp_map["__net_fin__"]
                cmp_map["__open__"] = cp_oc
                cmp_map["__close__"] = cp_cc
            else:
                cp_os = cp_from - timedelta(days=1)
                cp_pl = _period_movements(cp_from, cp_to, ["revenue", "expense", "contra-expense"])
                cp_np = 0.0
                for aid, (dr, cr) in cp_pl.items():
                    a = all_accts.get(aid)
                    if a is None: continue
                    if a.type == "expense":
                        cp_np += float(dr - cr if cr > dr else cr - dr)
                    else:
                        cp_np += float(cr - dr)
                cp_bs = _period_movements(cp_from, cp_to, ["asset", "liability", "equity"])
                cp_groups = {"operating": {}, "investing": {}, "financing": {}}
                for aid, (dr, cr) in cp_bs.items():
                    acct = all_accts.get(aid)
                    if acct is None: continue
                    act = acct.effective_cash_flow_activity() or "operating"
                    if act == "cash": continue
                    effect = float(cr - dr)
                    if effect:
                        head = l3_head(acct)
                        cp_groups[act][head] = cp_groups[act].get(head, 0) + effect
                # Net profit item
                cmp_map["Net Profit / (Loss) for the period"] = cp_np
                for head, v in sorted(cp_groups["operating"].items()):
                    label = f"(Increase) / Decrease in {head}" if v < 0 else f"Decrease / (Increase) in {head}"
                    cmp_map[label] = v
                for head, v in sorted(cp_groups["investing"].items()):
                    cmp_map[f"Movement in {head}"] = v
                for head, v in sorted(cp_groups["financing"].items()):
                    cmp_map[f"Movement in {head}"] = v
                cp_net_op = cp_np + sum(cp_groups["operating"].values())
                cp_net_inv = sum(cp_groups["investing"].values())
                cp_net_fin = sum(cp_groups["financing"].values())
                cmp_map["__net_op__"] = cp_net_op
                cmp_map["__net_inv__"] = cp_net_inv
                cmp_map["__net_fin__"] = cp_net_fin
                cmp_map["__net_chg__"] = cp_net_op + cp_net_inv + cp_net_fin
                cmp_map["__open__"] = cash_balance(cp_os)
                cmp_map["__close__"] = cash_balance(cp_to)
            comp_item_maps.append(cmp_map)

        # Annotate items with comp_vals
        def annotate_list(items, default=0):
            new_list = []
            for name, val in items:
                comps = [m.get(name, default) for m in comp_item_maps]
                new_list.append((name, val, comps))
            return new_list

        op_items = annotate_list(op_items)
        inv_items = annotate_list(inv_items)
        fin_items = annotate_list(fin_items)

    fmt = request.args.get("format")
    # Count the comparative maps actually built, not the periods that happen to
    # be in the query string. Switching the filter back from comparative leaves
    # comp_period_ids populated with comp_mode cleared, which opened unlabelled
    # columns of zeros in Excel and of whitespace in the PDF.
    n_comp = len(comp_item_maps)
    if fmt == "excel":
        col_count = 2 + n_comp
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cash Flow"
        first_row = _write_sheet_heading(
            ws, col_count, f"Cash Flow Statement ({method.title()} Method)",
            from_date=from_date, to_date=to_date)

        def write_section_excel(sr, title, items, total_label, total_val, comp_total=None):
            ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=col_count)
            ws.cell(row=sr, column=1, value=title).font = Font(bold=True, size=12)
            r = sr + 1
            for entry in items:
                if isinstance(entry, tuple) and len(entry) == 3:
                    name, val, comps = entry
                else:
                    name, val = entry
                    comps = [None] * n_comp
                ws.cell(row=r, column=1, value=name).font = DATA_FONT
                cc = ws.cell(row=r, column=2, value=val)
                cc.font = DATA_FONT; cc.alignment = RIGHT
                for ci, cv in enumerate(comps):
                    c = ws.cell(row=r, column=3+ci, value=cv or 0)
                    c.font = DATA_FONT; c.alignment = RIGHT
                r += 1
            ws.cell(row=r, column=1, value=total_label).font = BOLD_FONT
            cc = ws.cell(row=r, column=2, value=total_val)
            cc.font = BOLD_FONT; cc.alignment = RIGHT
            if comp_total is not None:
                for ci, cv in enumerate(comp_total):
                    c = ws.cell(row=r, column=3+ci, value=cv or 0)
                    c.font = BOLD_FONT; c.alignment = RIGHT
            return r + 2

        comp_net_op = [m["__net_op__"] for m in comp_item_maps] if comp_item_maps else []
        comp_net_inv = [m["__net_inv__"] for m in comp_item_maps] if comp_item_maps else []
        comp_net_fin = [m["__net_fin__"] for m in comp_item_maps] if comp_item_maps else []
        nr = write_section_excel(first_row, "OPERATING ACTIVITIES", op_items,
                                 "Net cash from operating activities", net_operating, comp_net_op)
        nr = write_section_excel(nr, "INVESTING ACTIVITIES", inv_items,
                                 "Net cash from investing activities", net_investing, comp_net_inv)
        nr = write_section_excel(nr, "FINANCING ACTIVITIES", fin_items,
                                 "Net cash from financing activities", net_financing, comp_net_fin)
        comp_net_chg = [m["__net_chg__"] for m in comp_item_maps] if comp_item_maps else []
        comp_open = [m["__open__"] for m in comp_item_maps] if comp_item_maps else []
        comp_close = [m["__close__"] for m in comp_item_maps] if comp_item_maps else []
        for label, val, comps in [("NET CHANGE IN CASH", net_change, comp_net_chg),
                                   ("Opening cash & equivalents", opening_cash, comp_open),
                                   ("Closing cash & equivalents", closing_cash, comp_close)]:
            ws.cell(row=nr, column=1, value=label).font = BOLD_FONT
            cc = ws.cell(row=nr, column=2, value=val)
            cc.font = BOLD_FONT; cc.alignment = RIGHT
            for ci, cv in enumerate(comps):
                c = ws.cell(row=nr, column=3+ci, value=cv or 0)
                c.font = BOLD_FONT; c.alignment = RIGHT
            nr += 1
        _finish_sheet(ws, col_count, first_row=first_row)
        out = BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, as_attachment=True, download_name="cash_flow.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if fmt == "pdf":
        pdf_headers = ["Item", _col_label(to_date)]
        for cp in comp_periods:
            pdf_headers.append(_col_label(cp.end_date))
        def _comp_str(v):
            return format_amount(v) if v is not None else ""
        def _comp_vals(maps, key, n):
            if maps:
                return [_comp_str(m.get(key, 0)) for m in maps]
            return [""] * n

        pdf_data, kinds = [], []

        def _add(row, kind):
            pdf_data.append(row)
            kinds.append(kind)

        def _activity(title, items, total_label, total_val, comp_key):
            _add([title, ""] + [""] * n_comp, "section")
            for entry in items:
                name, val = entry[0], entry[1]
                comps = entry[2] if len(entry) == 3 else [None] * n_comp
                _add([name, format_amount(val)] + [_comp_str(cv) for cv in comps],
                     "account")
            _add([total_label, format_amount(total_val)] +
                 _comp_vals(comp_item_maps, comp_key, n_comp), "total")
            _add(["", ""] + [""] * n_comp, "spacer")

        _activity("OPERATING ACTIVITIES", op_items,
                  "Net cash from operating activities", net_operating, "__net_op__")
        _activity("INVESTING ACTIVITIES", inv_items,
                  "Net cash from investing activities", net_investing, "__net_inv__")
        _activity("FINANCING ACTIVITIES", fin_items,
                  "Net cash from financing activities", net_financing, "__net_fin__")
        _add(["NET CHANGE IN CASH", format_amount(net_change)] +
             _comp_vals(comp_item_maps, "__net_chg__", n_comp), "grand")
        _add(["Opening cash & equivalents", format_amount(opening_cash)] +
             _comp_vals(comp_item_maps, "__open__", n_comp), "plain")
        _add(["Closing cash & equivalents", format_amount(closing_cash)] +
             _comp_vals(comp_item_maps, "__close__", n_comp), "total")
        pdf_out = _build_pdf(f"Cash Flow Statement ({method.title()} Method)",
                             pdf_headers, pdf_data,
                             subtitle=_period_line(from_date, to_date),
                             row_kinds=kinds, indent_col=0)
        return send_file(pdf_out, as_attachment=True, download_name="cash_flow.pdf",
                         mimetype="application/pdf")

    return render_template("finance/cash_flow.html", op_items=op_items,
                           inv_items=inv_items, fin_items=fin_items,
                           net_operating=net_operating, net_investing=net_investing,
                           net_financing=net_financing, net_change=net_change,
                           opening_cash=opening_cash, closing_cash=closing_cash,
                           cash_movement=cash_movement, method=method,
                           comp_item_maps=comp_item_maps,
                           from_date=from_date, to_date=to_date,
                           periods=periods, selected_period_id=selected_period_id,
                           filter_mode=filter_mode, from_str=from_str, to_str=to_str,
                           comp_mode=comp_mode, comp_periods=comp_periods, comp_period_ids_str=comp_period_ids_str,
                           now=datetime.utcnow())





# ═══════════════════════════════════════════════
# 8. 13-WEEK CASH FLOW (TWCF)
# ═══════════════════════════════════════════════


def _twcf_default_start():
    """The standard rolling start of a 13-week window: this week's Monday."""
    today = date.today()
    return today - timedelta(days=today.weekday())


def _twcf_week_label(wk):
    return f"{wk['start']:%d %b} – {wk['end']:%d %b}"


@finance_bp.route("/twcf", methods=["GET", "POST"])
@login_required
def twcf():
    """13-Week Cash Flow forecast: auto AR/AP schedule + user lines + actuals.

    GET renders the matrix; POST manages the forecast lines (add/edit/delete)
    and bounces back to the same window. ``format=excel|pdf`` exports the
    matrix exactly as shown.
    """
    from shared import twcf_reports as tw

    if request.method == "POST":
        action = request.form.get("action", "")
        start_arg = (request.form.get("start") or request.args.get("start")
                     or "").strip()
        back = url_for("finance.twcf", start=start_arg)

        if action in ("add", "edit"):
            line_id = request.form.get("line_id", type=int)
            line = scoped_get(TwcfLine, line_id) if line_id else None
            if action == "edit" and line is None:
                flash("Forecast line not found.", "danger")
                return redirect(back)

            direction = request.form.get("direction", "")
            category = request.form.get("category", "")
            description = (request.form.get("description") or "").strip()
            amount = request.form.get("amount", "").strip()
            start_raw = (request.form.get("start_date") or "").strip()
            frequency = request.form.get("frequency", "oneoff") or "oneoff"
            try:
                amount_val = Decimal(amount)
                start_val = _parse_date(start_raw)
            except (InvalidOperation, TypeError):
                amount_val, start_val = None, None

            allowed_cats = (TWCF_USER_IN_CATEGORIES if direction == TWCF_IN
                            else TWCF_USER_OUT_CATEGORIES)
            if direction not in (TWCF_IN, TWCF_OUT):
                flash("Choose whether this line is a receipt or a payment.",
                      "danger")
            elif category not in allowed_cats:
                flash("Choose a valid category for that direction.", "danger")
            elif not description:
                flash("Description is required.", "danger")
            elif amount_val is None or amount_val <= 0:
                flash("Amount must be a positive number.", "danger")
            elif start_val is None:
                flash("Start date is required.", "danger")
            elif frequency not in TWCF_FREQUENCIES:
                flash("Unknown frequency.", "danger")
            else:
                try:
                    day_of_week = max(0, min(6, int(request.form.get(
                        "day_of_week", 0) or 0)))
                    day_of_month = max(1, min(31, int(request.form.get(
                        "day_of_month", 1) or 1)))
                    month = max(1, min(12, int(request.form.get(
                        "month", 1) or 1)))
                except (TypeError, ValueError):
                    day_of_week, day_of_month, month = 0, 1, 1

                if action == "add":
                    db.session.add(TwcfLine(
                        direction=direction, category=category,
                        description=description, amount=amount_val,
                        start_date=start_val, frequency=frequency,
                        day_of_week=day_of_week, day_of_month=day_of_month,
                        month=month))
                    flash(f"Forecast line added: {description}", "success")
                else:
                    line.direction = direction
                    line.category = category
                    line.description = description
                    line.amount = amount_val
                    line.start_date = start_val
                    line.frequency = frequency
                    line.day_of_week = day_of_week
                    line.day_of_month = day_of_month
                    line.month = month
                    flash(f"Forecast line updated: {description}", "success")
                db.session.commit()
            return redirect(back)

        if action == "delete":
            line = scoped_get(TwcfLine, request.form.get("line_id", type=int))
            if line is not None:
                db.session.delete(line)
                db.session.commit()
                flash(f"Forecast line deleted: {line.description}", "success")
            return redirect(back)

    start_str = (request.args.get("start") or "").strip()
    start = _parse_date(start_str) if start_str else _twcf_default_start()
    matrix = tw.build_matrix(start)
    weeks = matrix["weeks"]
    lines = TwcfLine.query.order_by(TwcfLine.start_date, TwcfLine.id).all()
    edit_id = request.args.get("edit", type=int)
    editing = scoped_get(TwcfLine, edit_id) if edit_id else None

    fmt = request.args.get("format")

    if fmt == "excel":
        ncols = len(weeks) + 2  # Category + 13 weeks + Total
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "13 Week Cash Flow"
        first_row = _write_sheet_heading(
            ws, ncols, "13 Week Cash Flow (TWCF)",
            period=(f"Forecast window {_col_label(start)} to "
                    f"{_col_label(weeks[-1]['end'])}"))
        ws.cell(row=first_row, column=1, value="Category")
        for i, wk in enumerate(weeks):
            ws.cell(row=first_row, column=2 + i,
                    value=f"Wk {i + 1}\n{_twcf_week_label(wk)}")
        ws.cell(row=first_row, column=ncols, value="Total")
        for ci in range(1, ncols + 1):
            c = ws.cell(row=first_row, column=ci)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = THIN

        r = first_row + 1
        for row in matrix["rows"]:
            kind = row["kind"]
            if kind == "section":
                ws.merge_cells(start_row=r, start_column=1, end_row=r,
                               end_column=ncols)
                c = ws.cell(row=r, column=1, value=row["label"])
                c.font = Font(bold=True, size=11, color="334155")
                c.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                for ci in range(1, ncols + 1):
                    ws.cell(row=r, column=ci).border = THIN
                r += 1
                continue
            label_cell = ws.cell(row=r, column=1, value=row["label"])
            values = row["values"]
            for i in range(len(weeks)):
                v = values[i] if i < len(values) else None
                if v is not None:
                    cc = ws.cell(row=r, column=2 + i, value=v)
                    cc.number_format = MONEY_FMT
            total = row.get("total")
            if total is not None:
                tc = ws.cell(row=r, column=ncols, value=total)
                tc.number_format = MONEY_FMT
            if kind == "grand":
                for ci in range(1, ncols + 1):
                    c = ws.cell(row=r, column=ci)
                    c.font = Font(bold=True, size=11, color="FFFFFF")
                    c.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            elif kind in ("total", "net", "headroom"):
                for ci in range(1, ncols + 1):
                    c = ws.cell(row=r, column=ci)
                    c.font = BOLD_FONT
                    c.fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
                if kind == "net":
                    for ci in range(2, ncols + 1):
                        ws.cell(row=r, column=ci).fill = PatternFill(
                            start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            elif kind == "variance":
                for ci in range(1, ncols + 1):
                    c = ws.cell(row=r, column=ci)
                    c.font = Font(italic=True, size=10)
            elif kind == "opening":
                label_cell.font = Font(italic=True, size=10, bold=True)
            else:
                for ci in range(1, ncols + 1):
                    ws.cell(row=r, column=ci).font = DATA_FONT
            for ci in range(1, ncols + 1):
                c = ws.cell(row=r, column=ci)
                c.border = THIN
                if c.alignment is None or not c.alignment.horizontal:
                    c.alignment = RIGHT if ci > 1 else LEFT_ALIGN
            r += 1

        ws.freeze_panes = f"B{first_row + 1}"
        ws.column_dimensions["A"].width = 36
        for i in range(len(weeks)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(
                2 + i)].width = 13
        ws.column_dimensions[openpyxl.utils.get_column_letter(ncols)].width = 14
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, as_attachment=True,
                         download_name="13_week_cash_flow.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if fmt == "pdf":
        pdf_headers = (["Category"]
                       + [f"Wk {i + 1}\n{_twcf_week_label(wk)}"
                          for i, wk in enumerate(weeks)]
                       + ["Total"])
        pdf_data, kinds = [], []
        for row in matrix["rows"]:
            if row["kind"] == "section":
                pdf_data.append([row["label"]] + [""] * (len(weeks) + 1))
                kinds.append("section")
                continue
            cells = [row["label"]]
            cells += [format_amount(v) if v is not None else ""
                      for v in row["values"]]
            total = row.get("total")
            cells.append(format_amount(total) if total is not None else "")
            kind = row["kind"]
            pdf_kind = {"opening": "account", "net": "subtotal",
                        "variance": "plain", "headroom": "subtotal"}.get(
                kind, kind)
            pdf_data.append(cells)
            kinds.append(pdf_kind)
        subtitle = (f"Forecast window {_col_label(start)} to "
                    f"{_col_label(weeks[-1]['end'])}")
        pdf_out = _build_pdf("13 Week Cash Flow (TWCF)", pdf_headers,
                             pdf_data, subtitle=subtitle, row_kinds=kinds,
                             indent_col=0)
        return send_file(pdf_out, as_attachment=True,
                         download_name="13_week_cash_flow.pdf",
                         mimetype="application/pdf")

    return render_template(
        "finance/twcf.html",
        matrix=matrix, weeks=weeks, start=start,
        lines=lines, editing=editing,
        in_categories=TWCF_USER_IN_CATEGORIES,
        out_categories=TWCF_USER_OUT_CATEGORIES,
        frequencies=TWCF_FREQUENCIES,
        out_cat_labels={**TWCF_IN_CATEGORIES, **TWCF_OUT_CATEGORIES},
        frequency_labels={"oneoff": "One-off", "weekly": "Weekly",
                          "monthly": "Monthly", "quarterly": "Quarterly",
                          "yearly": "Yearly"},
        now=datetime.utcnow())
