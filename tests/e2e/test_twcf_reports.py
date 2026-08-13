"""13-Week Cash Flow (TWCF) — HTTP-level end-to-end coverage.

Runs against the real app with a throwaway database (same pattern as
test_executive_reports.py): login via session, seed journal postings
directly, then drive the report over HTTP — matrix rendering, actuals and
variance for elapsed weeks, forecast-line CRUD, the cash floor, and the
Excel/PDF exports.
"""

import io
import os
import tempfile
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from decimal import Decimal

import pytest

_TMP_DB = os.path.join(tempfile.gettempdir(), "erp_twcf_e2e.db")
if os.path.exists(_TMP_DB):
    os.remove(_TMP_DB)
os.environ["DATABASE_URL"] = "sqlite:///" + _TMP_DB.replace("\\", "/")

from app import app as flask_app  # noqa: E402
from shared.extensions import db  # noqa: E402


@contextmanager
def _ctx(book):
    """An app context with the active company selected (tenancy fails closed)."""
    from shared.tenancy import set_current_company
    with flask_app.app_context():
        set_current_company(book["company_id"])
        yield


@pytest.fixture(scope="module")
def client():
    c = flask_app.test_client()
    c.get("/")  # lazy create_all + migrate + seed
    from hr_app.models.user import User
    with flask_app.app_context():
        user = (User.query.filter(User.email.ilike("admin%")).first()
                or User.query.first())
        assert user is not None, "seed produced no users"
        uid = user.id
    with c.session_transaction() as sess:
        sess["_user_id"] = str(uid)
        sess["_fresh"] = True
    yield c
    with flask_app.app_context():
        db.session.remove()


@pytest.fixture(scope="module")
def book(client):
    """Cash, receivable and payable accounts from the seeded fixed chart."""
    from shared.models.company import Company
    from shared.models.ledger import ChartOfAccount
    from shared.tenancy import set_current_company, unscoped

    with flask_app.app_context():
        with unscoped():
            company = Company.query.order_by(Company.id).first()
        assert company is not None
        set_current_company(company.id)
        cash = ChartOfAccount.query.filter_by(code="1-01-01-01-0001").first()
        ar = ChartOfAccount.query.filter_by(code="1-01-02-01-0001").first()
        ap = ChartOfAccount.query.filter_by(code="2-01-01-01-0001").first()
        assert cash and ar and ap, "seeded chart missing role accounts"
        return {"company_id": company.id, "cash": cash.id,
                "ar": ar.id, "ap": ap.id}


def _post(book, account_id, debit=0, credit=0, when=None):
    from shared.models.ledger import JournalEntry, JournalLine
    with _ctx(book):
        from hr_app.models.user import User
        uid = User.query.first().id
        entry = JournalEntry(voucher_type="JV", voucher_id=0,
                             voucher_number=f"TWCFE2E-{int(datetime.utcnow().timestamp())}",
                             description="twcf e2e",
                             entry_date=when or datetime.utcnow(),
                             created_by=uid, is_posted=True)
        db.session.add(entry)
        db.session.commit()
        db.session.add(JournalLine(journal_entry_id=entry.id,
                                   account_id=account_id,
                                   debit=Decimal(str(debit)),
                                   credit=Decimal(str(credit))))
        db.session.commit()
        return entry


def _clear_journals(book):
    from shared.models.ledger import JournalEntry, JournalLine
    with _ctx(book):
        JournalLine.query.delete()
        JournalEntry.query.delete()
        db.session.commit()


def test_matrix_renders_full_twcf_structure(client, book):
    start = date(2026, 8, 3)
    r = client.get(f"/finance/twcf?start={start:%Y-%m-%d}")
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "13-Week Cash Flow" in html
    assert "Forecast window" in html
    assert "RECEIPTS" in html and "DISBURSEMENTS" in html
    assert "Opening cash &amp; equivalents" in html
    assert "Collections from customers" in html
    assert "Payments to suppliers" in html
    assert "Total receipts" in html and "Total disbursements" in html
    assert "Net cash flow" in html
    assert "Closing cash &amp; equivalents" in html
    for i in range(1, 14):
        assert f"Wk {i}" in html
    assert "Forecast Lines" in html


def test_matrix_auto_rows_come_from_aged_ledger(client, book):
    _clear_journals(book)
    start = date(2026, 8, 3)
    # Customer owes 1400, posted 5 days before the window (current bucket).
    _post(book, book["ar"], debit=1400, when=datetime(2026, 7, 29, 12, 0))
    # Supplier to be paid 700, 20 days before (current bucket).
    _post(book, book["ap"], credit=700, when=datetime(2026, 7, 14, 12, 0))
    r = client.get(f"/finance/twcf?start={start:%Y-%m-%d}")
    html = r.get_data(as_text=True)
    # 1400/2 = 700 in week 1, 700/2 = 350 in week 1 for the payable.
    assert "700.00" in html and "350.00" in html


def test_past_weeks_show_actuals_and_variance(client, book):
    _clear_journals(book)
    start = date(2026, 8, 3)  # week 1 (Aug 3-9) is fully elapsed whenever
    # the suite runs after 9 Aug 2026.
    _post(book, book["ar"], debit=1400, when=datetime(2026, 7, 29, 12, 0))
    # Actual cash receipt of 900 in week 1 (Aug 5).
    _post(book, book["cash"], debit=900, when=datetime(2026, 8, 5, 12, 0))
    r = client.get(f"/finance/twcf?start={start:%Y-%m-%d}")
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "Variance (actual vs forecast)" in html
    # Forecast collections in week 1 = 700; actual net = 900 -> variance 200.
    assert "200.00" in html


def test_forecast_line_add_edit_delete_roundtrip(client, book):
    _clear_journals(book)
    from shared.models.twcf import TwcfLine
    start = date(2026, 8, 3)

    r = client.post("/finance/twcf", data={
        "action": "add", "start": f"{start:%Y-%m-%d}",
        "direction": "out", "category": "payroll", "description": "Salaries",
        "amount": "2500", "start_date": f"{start:%Y-%m-%d}",
        "frequency": "monthly", "day_of_month": "28",
    })
    assert r.status_code == 302
    with _ctx(book):
        line = TwcfLine.query.filter_by(description="Salaries").first()
        assert line is not None
        assert line.category == "payroll" and line.frequency == "monthly"
        assert line.day_of_month == 28
        line_id = line.id

    # The line shows up in the matrix under Salaries & wages.
    r = client.get(f"/finance/twcf?start={start:%Y-%m-%d}")
    html = r.get_data(as_text=True)
    assert "Salaries &amp; wages" in html
    assert "2,500.00" in html

    # Edit: change amount and day.
    r = client.post("/finance/twcf", data={
        "action": "edit", "start": f"{start:%Y-%m-%d}", "line_id": line_id,
        "direction": "out", "category": "payroll", "description": "Salaries",
        "amount": "3000", "start_date": f"{start:%Y-%m-%d}",
        "frequency": "monthly", "day_of_month": "1",
    })
    assert r.status_code == 302
    with _ctx(book):
        line = TwcfLine.query.get(line_id)
        assert line.amount == Decimal("3000.00")
        assert line.day_of_month == 1

    # Delete.
    r = client.post("/finance/twcf", data={
        "action": "delete", "start": f"{start:%Y-%m-%d}", "line_id": line_id})
    assert r.status_code == 302
    with _ctx(book):
        assert TwcfLine.query.get(line_id) is None


def test_forecast_line_validation_rejects_bad_input(client, book):
    start = date(2026, 8, 3)
    r = client.post("/finance/twcf", data={
        "action": "add", "start": f"{start:%Y-%m-%d}",
        "direction": "out", "category": "payroll", "description": "Salaries",
        "amount": "-5", "start_date": f"{start:%Y-%m-%d}",
        "frequency": "monthly", "day_of_month": "28",
    }, follow_redirects=True)
    assert r.status_code == 200
    assert "Amount must be a positive number" in r.get_data(as_text=True)


def test_cash_floor_enables_headroom_rows(client, book):
    _clear_journals(book)
    from shared.models.company_settings import ReportSettings
    start = date(2026, 8, 3)
    with _ctx(book):
        s = ReportSettings.get()
        s.twcf_cash_floor = Decimal("4000")
        db.session.commit()
    r = client.get(f"/finance/twcf?start={start:%Y-%m-%d}")
    html = r.get_data(as_text=True)
    assert "Minimum cash (floor)" in html
    assert "Headroom over floor" in html
    with _ctx(book):
        s = ReportSettings.get()  # fresh query: the first block's instance
        # is detached once its app context popped, and mutating it would
        # silently not persist
        s.twcf_cash_floor = None
        db.session.commit()
    r = client.get(f"/finance/twcf?start={start:%Y-%m-%d}")
    assert "Headroom over floor" not in r.get_data(as_text=True)


def test_excel_export_matches_screen_structure(client, book):
    _clear_journals(book)
    start = date(2026, 8, 3)
    _post(book, book["cash"], debit=10000, when=datetime(2026, 8, 1, 12, 0))
    r = client.get(f"/finance/twcf?start={start:%Y-%m-%d}&format=excel")
    assert r.status_code == 200
    assert r.mimetype == ("application/vnd.openxmlformats-officedocument"
                          ".spreadsheetml.sheet")
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    ws = wb["13 Week Cash Flow"]
    assert ws["A1"].value is not None  # company
    assert "13 Week Cash Flow (TWCF)" in str(ws["A2"].value)
    assert "Forecast window" in str(ws["A3"].value)
    # Header row: Category + 13 weeks + Total, wrapped two-line week labels.
    hdr = 5
    assert ws.cell(row=hdr, column=1).value == "Category"
    assert ws.cell(row=hdr, column=15).value == "Total"
    assert str(ws.cell(row=hdr, column=2).value).startswith("Wk 1")
    assert "\n" in str(ws.cell(row=hdr, column=2).value)
    # Opening cash figure is money-formatted.
    money = ws.cell(row=hdr + 1, column=2).number_format
    assert money and "#,##0" in money
    # Freeze keeps the label column and header rows pinned.
    assert ws.freeze_panes == "B6"


def test_pdf_export_builds_landscape_report(client, book):
    _clear_journals(book)
    start = date(2026, 8, 3)
    r = client.get(f"/finance/twcf?start={start:%Y-%m-%d}&format=pdf")
    assert r.status_code == 200
    assert r.mimetype == "application/pdf"
    from PyPDF2 import PdfReader
    pdf = PdfReader(io.BytesIO(r.data))
    assert len(pdf.pages) == 1
    text = pdf.pages[0].extract_text() or ""
    assert "13 Week Cash Flow (TWCF)" in text
    assert "Forecast window" in text
    # Landscape: width exceeds height.
    w, h = pdf.pages[0].mediabox.width, pdf.pages[0].mediabox.height
    assert float(w) > float(h)
    # The 13 weekly columns are all present.
    assert "Wk 1" in text and "Wk 13" in text


def test_start_date_defaults_to_current_week(client, book):
    r = client.get("/finance/twcf")
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    today = datetime.utcnow().date()
    monday = today - timedelta(days=today.weekday())
    assert f"{monday:%Y-%m-%d}" in html
