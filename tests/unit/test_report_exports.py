"""Finance report exports.

The rule these protect: what the screen shows is what the file contains, at a
size that fits the page, with money formatted as money. Every one of these
covers a defect that shipped — a report that printed off the edge of the paper,
or a spreadsheet full of raw floats — because nothing here was exercised.
"""

import io
import re
from datetime import date
from decimal import Decimal

import openpyxl
import pytest
from PyPDF2 import PdfReader

from finance_app.routes.reports import (
    MONEY_FMT, SHEET_FIRST_ROW, _build_excel_wb, _build_pdf, _cell_text,
    _col_label, _finish_sheet, _pl_comp_amount)


HEADERS = ["Date", "Voucher", "Account Code", "Account Name", "Description",
           "Debit", "Credit", "Balance"]


def _rows(n=3):
    return [["2026-07-%02d" % ((i % 28) + 1), "JV-202607-%04d" % i, "5-2-03-00704",
             "Selling and Distribution Expenses - Carriage Outward",
             "Being carriage paid on delivery to the customer site",
             12345678.90, 0.0, 98765432.10] for i in range(n)]


def _pdf(headers=None, rows=None, **kw):
    buf = _build_pdf("General Ledger", headers or HEADERS, rows or _rows(), **kw)
    return PdfReader(io.BytesIO(buf.getvalue()))


def _page_text(page):
    return page.extract_text() or ""


def _font_sizes(page):
    """Every font size the page actually draws with."""
    data = page.get_contents().get_data().decode("latin-1")
    return sorted({round(float(m), 2) for m in re.findall(r"/F\d+ ([\d.]+) Tf", data)})


# ─────────────────────────────────────────────
# PDF
# ─────────────────────────────────────────────

def test_every_page_is_numbered():
    """The callbacks were handed to the SimpleDocTemplate constructor, which
    takes arbitrary keywords and drops them, so no PDF was ever numbered."""
    r = _pdf(rows=_rows(120))
    assert len(r.pages) > 1, "expected a multi-page report"
    for i, page in enumerate(r.pages, 1):
        assert "Page %d" % i in _page_text(page)


def test_a_wide_report_goes_landscape():
    portrait = _pdf(headers=HEADERS[:4], rows=[r[:4] for r in _rows()])
    landscape = _pdf()
    pw, ph = portrait.pages[0].mediabox.width, portrait.pages[0].mediabox.height
    lw, lh = landscape.pages[0].mediabox.width, landscape.pages[0].mediabox.height
    assert float(pw) < float(ph), "few columns stay portrait"
    assert float(lw) > float(lh), "many columns turn landscape"


def test_the_type_shrinks_as_columns_are_added():
    """The fitted size only reached cells wrapped in a Paragraph; everything
    else kept reportlab's 10pt default and ran off the page."""
    def smallest(ncols):
        heads = ["Column Heading %d" % i for i in range(ncols)]
        rows = [["Carriage Outward Expenses"] + [1234567.89] * (ncols - 1)
                for _ in range(4)]
        return min(_font_sizes(_pdf(headers=heads, rows=rows).pages[0]))

    narrow, wide = smallest(4), smallest(16)
    assert wide < narrow, "a wider table has to set smaller type"
    assert wide >= 5.0, "but never below legibility"


def test_headings_are_not_broken_mid_word():
    """Columns budgeted less width than reportlab's own cell padding, so every
    one came out narrower than its widest word: Vouche/r, Balanc/e."""
    text = _page_text(_pdf().pages[0])
    for head in ("Voucher", "Balance", "Credit", "Description"):
        assert head in text, head
        assert head[:-1] + "\n" not in text, "%s broke mid-word" % head


def test_the_heading_row_is_drawn_bold():
    """Header cells are Paragraphs, which draw their own text — TableStyle's
    white bold heading never reached them, leaving black on dark blue."""
    page = _pdf().pages[0]
    fonts = {v.get_object().get("/BaseFont")
             for v in page["/Resources"]["/Font"].values()}
    assert "/Helvetica-Bold" in {str(f) for f in fonts}


def test_the_report_title_and_subtitle_survive():
    text = _page_text(_pdf(subtitle="01-Jul-2026 to 31-Jul-2026").pages[0])
    assert "General Ledger" in text
    assert "01-Jul-2026" in text


def test_an_empty_report_still_builds():
    """A period with no entries must produce a page, not an exception."""
    r = _pdf(rows=[])
    assert len(r.pages) == 1
    assert "General Ledger" in _page_text(r.pages[0])


# ─────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────

def _sheet(**kw):
    buf = _build_excel_wb("Trial Balance", ["Code", "Account Name", "Debit", "Credit"],
                          [["1-1-01", "Cash at Bank", 1234567.891, 0.0],
                           ["5-2-03", "Carriage Outward", 0.0, -9876.5]], **kw)
    return openpyxl.load_workbook(io.BytesIO(buf.getvalue())).active


# Headings occupy rows 1-3 and column headers sit on SHEET_FIRST_ROW, so the
# first row of figures is the one after that.
FIRST_DATA_ROW = SHEET_FIRST_ROW + 1


def test_money_cells_carry_a_money_format():
    """number_format was passed on one export path out of eight, so the rest
    wrote 1234.5 where the report says 1,234.50."""
    ws = _sheet()
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, min_col=3, max_col=4):
        for c in row:
            assert c.number_format == MONEY_FMT, (c.coordinate, c.number_format)


def test_text_cells_are_left_as_text():
    ws = _sheet()
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, min_col=1, max_col=2):
        for c in row:
            assert c.number_format == "General"


def test_money_stays_a_number_so_the_sheet_can_sum_it():
    ws = _sheet()
    assert isinstance(ws.cell(row=FIRST_DATA_ROW, column=3).value, (int, float, Decimal))


def test_a_sheet_leads_with_the_same_heading_block_as_the_screen():
    """Exports carried "Profit & Loss (2026-07-01 to 2027-06-30)" against a
    screen showing company, title and a worded period on three lines."""
    from datetime import date as _date
    ws = _sheet(from_date=_date(2026, 7, 1), to_date=_date(2027, 6, 30))
    assert ws.cell(row=2, column=1).value == "Trial Balance"
    assert ws.cell(row=3, column=1).value == (
        "For the period 01 July, 2026 to 30 June, 2027")
    assert ws.cell(row=SHEET_FIRST_ROW, column=1).value == "Code", "headers follow"


def test_columns_are_wide_enough_for_the_formatted_figure():
    """Sizing from str(1234567.891) budgets eleven characters for the thirteen
    that get drawn, and Excel shows ###### instead of a number."""
    ws = _sheet()
    width = ws.column_dimensions["C"].width
    assert width >= len("1,234,567.89"), width


@pytest.mark.parametrize("value,expected", [
    (1234567.891, "1,234,567.89"),
    # Brackets, not a minus — and a character wider, which is the reason the
    # width has to be measured through the formatter rather than str().
    (-9876.5, "(9,876.50)"),
    (0, "0.00"),
    (None, ""),
    ("Cash at Bank", "Cash at Bank"),
    (True, "True"),
])
def test_cell_text_reports_what_excel_will_show(value, expected):
    assert _cell_text(value) == expected


def test_finish_sheet_formats_a_hand_built_sheet():
    """The export routes that build their own sheet go through the same pass,
    rather than each carrying a copy that forgot the number format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=3, column=1, value="Account")
    ws.cell(row=3, column=2, value="Amount")
    ws.cell(row=4, column=1, value="Carriage Outward")
    ws.cell(row=4, column=2, value=1234567.891)
    _finish_sheet(ws, 2)
    assert ws.cell(row=4, column=2).number_format == MONEY_FMT
    assert ws.column_dimensions["B"].width >= len("1,234,567.89")


def test_finish_sheet_leaves_a_deliberate_format_alone():
    wb = openpyxl.Workbook()
    ws = wb.active
    c = ws.cell(row=4, column=1, value=0.155)
    c.number_format = "0.00%"
    _finish_sheet(ws, 1)
    assert c.number_format == "0.00%", "an explicit format is not overwritten"


# ─────────────────────────────────────────────
# Comparative columns
# ─────────────────────────────────────────────

def test_column_label_is_the_date_the_figures_are_stated_at():
    """Screen and exports label a column the same way or the export does not
    match the report it came from."""
    assert _col_label(date(2027, 6, 30)) == "30 June, 2027"
    assert _col_label(None) == "Amount"


def _lookup(accounts=None, totals=None, subtotals=None):
    return (accounts or {}, totals or {}, subtotals or {})


def test_a_comparative_section_total_reads_that_section_only():
    """The reported defect: Total Sales showed -100,000 in the comparative
    column while the only sales account under it showed 0.00, because the
    total summed every account in the period — it was the period's net
    profit sitting under a Sales heading."""
    lk = _lookup(accounts={"4-01-01-01-0001": 0.0, "5-01-01-01-0001": -100000.0},
                 totals={"sales": 0.0, "admin": -100000.0})
    row = {"kind": "total", "label": "Total Sales", "section": "sales"}
    assert _pl_comp_amount(row, lk) == 0.0


def test_a_comparative_subtotal_reads_its_own_running_figure():
    lk = _lookup(subtotals={"net_sales": 630000.0, "net_profit": 90000.0})
    assert _pl_comp_amount(
        {"kind": "subtotal", "label": "Net Sales", "subtotal": "net_sales"}, lk) == 630000.0
    assert _pl_comp_amount(
        {"kind": "subtotal", "label": "Net Profit", "subtotal": "net_profit"}, lk) == 90000.0


def test_an_account_missing_from_a_comparative_period_reads_zero():
    lk = _lookup(accounts={"4-01-01-01-0001": 630000.0})
    row = {"kind": "account", "code": "4-01-01-01-0002", "section": "sales"}
    assert _pl_comp_amount(row, lk) == 0.0


def test_an_others_row_takes_whatever_the_listed_accounts_do_not():
    """Which accounts collapse into "Others" differs period to period, so the
    row has to be the section's remainder or the column stops adding up."""
    lk = _lookup(accounts={"A": 100.0, "B": 50.0, "C": 25.0},
                 totals={"admin": 175.0})
    row = {"kind": "account", "code": "", "name": "Others (1 accounts)",
           "section": "admin", "shown_codes": ["A", "B"]}
    assert _pl_comp_amount(row, lk) == 25.0


# ─────────────────────────────────────────────
# Statement styling
# ─────────────────────────────────────────────

STMT_HEADERS = ["Code", "Account / Section", "31 December, 2027"]
STMT_ROWS = [
    ["SALES", "", ""],
    ["4-01-01-01-0001", "Sales General", "630,000.00"],
    ["", "Total Sales", "630,000.00"],
    ["", "Net Sales", "(100,000.00)"],
    ["", "", ""],
]
STMT_KINDS = ["section", "account", "total", "subtotal", "spacer"]


def _fills(page):
    """Every fill colour the page paints, as reportlab writes them."""
    data = page.get_contents().get_data().decode("latin-1")
    return set(re.findall(r"([\d.]+ [\d.]+ [\d.]+) rg", data))


def _strokes(page):
    data = page.get_contents().get_data().decode("latin-1")
    return set(re.findall(r"([\d.]+ [\d.]+ [\d.]+) RG", data))


def _statement_pdf(**kw):
    buf = _build_pdf("Profit &amp; Loss Statement", STMT_HEADERS, STMT_ROWS,
                     row_kinds=STMT_KINDS, indent_col=1, mono_col=0, **kw)
    return PdfReader(io.BytesIO(buf.getvalue()))


def test_a_statement_pdf_bands_its_sections():
    """Section headings rendered as ordinary striped rows, indistinguishable
    from the accounts under them."""
    page = _statement_pdf().pages[0]
    assert ".945098 .960784 .976471" in _fills(page), "#F1F5F9 section band"


def test_totals_and_profit_lines_are_ruled_off():
    page = _statement_pdf().pages[0]
    strokes = _strokes(page)
    assert ".580392 .639216 .721569" in strokes, "#94A3B8 rule above a total"
    assert ".117647 .160784 .231373" in strokes, "#1E293B rule around a profit line"


def test_the_header_matches_the_screen_not_the_old_export_blue():
    page = _statement_pdf().pages[0]
    fills = _fills(page)
    assert ".117647 .160784 .231373" in fills, "#1E293B slate, as on screen"
    assert ".121569 .305882 .47451" not in fills, "#1F4E79 was the export's own blue"


def test_a_negative_profit_line_is_flagged_red():
    page = _statement_pdf().pages[0]
    assert ".933333 .933333 .933333" not in _fills(page)
    assert ".996078 .94902 .94902" in _fills(page), "#FEF2F2 behind a loss"


def test_the_statement_still_reads_in_order():
    text = _statement_pdf().pages[0].extract_text()
    for label in ["SALES", "Sales General", "Total Sales", "Net Sales"]:
        assert label in text
    assert text.index("SALES") < text.index("Total Sales")


def test_a_report_without_row_kinds_still_renders():
    """The ledger and any other caller keep working while they are converted."""
    r = _pdf()
    assert len(r.pages) >= 1
    assert "Account Code" in _page_text(r.pages[0])
